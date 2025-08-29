# -*- coding: utf-8 -*-
"""
Scheduler EXACTO del legacy Streamlit - Replicación 1:1
Contiene toda la lógica original para generar los mismos resultados
"""
import numpy as np
import time
import json
import hashlib
import gc
from itertools import combinations, permutations
from io import BytesIO, StringIO
import tempfile
import csv

try:
    import pulp as pl
    PULP_AVAILABLE = True
except ImportError:
    PULP_AVAILABLE = False

try:
    import psutil
except ImportError:
    psutil = None

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns
except ImportError:
    matplotlib = None
    plt = None
    sns = None

print(f"[LEGACY_SCHEDULER] PuLP disponible: {PULP_AVAILABLE}")

# Configuración EXACTA del legacy Streamlit
LEGACY_CONFIG = {
    "solver_time": 240,  # EXACTO del legacy
    "solver_msg": 1,
    "TARGET_COVERAGE": 98.0,
    "agent_limit_factor": 30,  # JEAN default
    "excess_penalty": 5.0,     # JEAN default
    "peak_bonus": 2.0,         # JEAN default
    "critical_bonus": 2.5,     # JEAN default
    "iterations": 30,
    "solver_threads": 4,
    "break_from_start": 2.5,
    "break_from_end": 2.5,
    "use_ft": True,
    "use_pt": True,
    "allow_8h": True,
    "allow_10h8": False,
    "allow_pt_4h": True,
    "allow_pt_6h": True,
    "allow_pt_5h": False,
}

def merge_legacy_config(cfg=None):
    """Combina configuración legacy con overrides."""
    merged = LEGACY_CONFIG.copy()
    if cfg:
        merged.update({k: v for k, v in cfg.items() if v is not None})
    return merged

def monitor_memory_usage():
    """Return current memory usage percentage."""
    if psutil is None:
        return 0.0
    return psutil.virtual_memory().percent

def adaptive_chunk_size(base=5000):
    """Adjust chunk size based on memory usage."""
    if psutil is None:
        return base
    usage = monitor_memory_usage()
    if usage > 80:
        return max(1000, base // 4)
    if usage > 60:
        return max(2000, base // 2)
    return base

def emergency_cleanup(threshold=85.0):
    """Trigger gc.collect if usage exceeds threshold."""
    if psutil is None:
        return False
    if monitor_memory_usage() >= threshold:
        gc.collect()
        return True
    return False

def score_pattern(pattern: np.ndarray, demand_matrix: np.ndarray) -> int:
    """Quick heuristic score to sort patterns before solving."""
    dm = demand_matrix.flatten()
    pat = pattern.astype(int)
    lim = min(len(dm), len(pat))
    return int(np.minimum(pat[:lim], dm[:lim]).sum())

def analyze_demand_matrix_legacy(matrix: np.ndarray) -> dict:
    """Análisis EXACTO del legacy Streamlit."""
    daily_demand = matrix.sum(axis=1)
    hourly_demand = matrix.sum(axis=0)
    active_days = [d for d in range(7) if daily_demand[d] > 0]
    inactive_days = [d for d in range(7) if daily_demand[d] == 0]
    working_days = len(active_days)
    active_hours = np.where(hourly_demand > 0)[0]
    first_hour = int(active_hours.min()) if active_hours.size else 8
    last_hour = int(active_hours.max()) if active_hours.size else 20
    operating_hours = last_hour - first_hour + 1
    peak_demand = float(matrix.max()) if matrix.size else 0.0
    avg_demand = float(matrix[active_days].mean()) if active_days else 0.0
    
    # Análisis crítico EXACTO del legacy
    daily_totals = matrix.sum(axis=1)
    hourly_totals = matrix.sum(axis=0)
    critical_days = (
        np.argsort(daily_totals)[-2:] if daily_totals.size > 1 else [int(np.argmax(daily_totals))]
    )
    peak_threshold = (
        np.percentile(hourly_totals[hourly_totals > 0], 75)
        if np.any(hourly_totals > 0)
        else 0
    )
    peak_hours = np.where(hourly_totals >= peak_threshold)[0]
    
    return {
        "daily_demand": daily_demand,
        "hourly_demand": hourly_demand,
        "active_days": active_days,
        "inactive_days": inactive_days,
        "working_days": working_days,
        "first_hour": first_hour,
        "last_hour": last_hour,
        "operating_hours": operating_hours,
        "peak_demand": peak_demand,
        "average_demand": avg_demand,
        "critical_days": critical_days,
        "peak_hours": peak_hours,
    }

def generate_weekly_pattern_legacy(start_hour, duration, working_days, dso_day=None, break_len=1, *, cfg=None):
    """Genera patrón semanal EXACTO del legacy Streamlit."""
    cfg = merge_legacy_config(cfg)
    break_from_start = cfg["break_from_start"]
    break_from_end = cfg["break_from_end"]
    pattern = np.zeros((7, 24), dtype=np.int8)
    
    for day in working_days:
        if day != dso_day:
            for h in range(duration):
                t = start_hour + h
                d_off, idx = divmod(int(t), 24)
                pattern[(day + d_off) % 7, idx] = 1
            
            # Break EXACTO del legacy
            break_start_idx = start_hour + break_from_start
            break_end_idx = start_hour + duration - break_from_end
            if int(break_start_idx) < int(break_end_idx):
                break_hour = int(break_start_idx) + (int(break_end_idx) - int(break_start_idx)) // 2
            else:
                break_hour = int(break_start_idx)
            
            for b in range(int(break_len)):
                t = break_hour + b
                d_off, idx = divmod(int(t), 24)
                pattern[(day + d_off) % 7, idx] = 0
    
    return pattern.flatten()

def generate_weekly_pattern_10h8_legacy(start_hour, working_days, eight_hour_day, break_len=1, *, cfg=None):
    """Genera patrón 10h+8h EXACTO del legacy."""
    cfg = merge_legacy_config(cfg)
    break_from_start = cfg["break_from_start"]
    break_from_end = cfg["break_from_end"]
    pattern = np.zeros((7, 24), dtype=np.int8)
    
    for day in working_days:
        duration = 8 if day == eight_hour_day else 10
        for h in range(duration):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
        
        break_start_idx = start_hour + break_from_start
        break_end_idx = start_hour + duration - break_from_end
        if int(break_start_idx) < int(break_end_idx):
            break_hour = int(break_start_idx) + (int(break_end_idx) - int(break_start_idx)) // 2
        else:
            break_hour = int(break_start_idx)
        
        for b in range(int(break_len)):
            t = break_hour + b
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 0
    
    return pattern.flatten()

def generate_weekly_pattern_simple_legacy(start_hour, duration, working_days):
    """Patrón simple sin break EXACTO del legacy."""
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        for h in range(duration):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
    return pattern.flatten()

def generate_weekly_pattern_pt5_legacy(start_hour, working_days):
    """Patrón PT5 EXACTO del legacy (5h en 4 días, 4h en 1 día)."""
    pattern = np.zeros((7, 24), dtype=np.int8)
    if not working_days:
        return pattern.flatten()
    
    four_hour_day = working_days[-1]
    for day in working_days:
        hours = 4 if day == four_hour_day else 5
        for h in range(hours):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
    
    return pattern.flatten()

def generate_shifts_coverage_legacy(*, cfg=None):
    """Generación de patrones EXACTA del legacy Streamlit."""
    cfg = merge_legacy_config(cfg)
    use_ft = cfg["use_ft"]
    use_pt = cfg["use_pt"]
    allow_8h = cfg["allow_8h"]
    allow_10h8 = cfg["allow_10h8"]
    allow_pt_4h = cfg["allow_pt_4h"]
    allow_pt_6h = cfg["allow_pt_6h"]
    allow_pt_5h = cfg["allow_pt_5h"]
    
    # Días activos EXACTO del legacy
    active_days = list(range(7))  # Todos los días por defecto
    
    shifts_coverage = {}
    seen_patterns = set()
    step = 0.5
    start_hours = [h for h in np.arange(0, 24, step) if h <= 23.5]
    
    print(f"[LEGACY_GEN] Generando patrones: FT={use_ft}, PT={use_pt}")
    
    for start_hour in start_hours:
        # Full Time 8 horas - 6 días de trabajo EXACTO del legacy
        if use_ft and allow_8h:
            for dso_day in active_days:
                working_days = [d for d in active_days if d != dso_day][:6]
                if len(working_days) >= 6 and 8 * len(working_days) <= 48:
                    pattern = generate_weekly_pattern_legacy(start_hour, 8, working_days, dso_day, cfg=cfg)
                    key = pattern.tobytes()
                    if key not in seen_patterns:
                        seen_patterns.add(key)
                        name = f"FT8_{start_hour:04.1f}_DSO{dso_day}"
                        shifts_coverage[name] = pattern
        
        # Full Time 10h + 8h - 5 días EXACTO del legacy
        if use_ft and allow_10h8:
            for dso_day in active_days:
                working_days = [d for d in active_days if d != dso_day][:5]
                if len(working_days) >= 5:
                    for eight_day in working_days:
                        pattern = generate_weekly_pattern_10h8_legacy(start_hour, working_days, eight_day, cfg=cfg)
                        key = pattern.tobytes()
                        if key not in seen_patterns:
                            seen_patterns.add(key)
                            name = f"FT10p8_{start_hour:04.1f}_DSO{dso_day}_8{eight_day}"
                            shifts_coverage[name] = pattern
        
        # Part Time patterns EXACTO del legacy
        if use_pt:
            # 4 horas - múltiples combinaciones
            if allow_pt_4h:
                for num_days in [4, 5, 6]:
                    if num_days <= len(active_days) and 4 * num_days <= 24:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple_legacy(start_hour, 4, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT4_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
            
            # 6 horas - 4 días (24h/week)
            if allow_pt_6h:
                for num_days in [4]:
                    if num_days <= len(active_days) and 6 * num_days <= 24:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple_legacy(start_hour, 6, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT6_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
            
            # 5 horas - 5 días (~25h/week)
            if allow_pt_5h:
                for num_days in [5]:
                    if num_days <= len(active_days) and 5 * num_days <= 25:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_pt5_legacy(start_hour, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT5_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
    
    print(f"[LEGACY_GEN] Generados {len(shifts_coverage)} patrones únicos")
    return shifts_coverage

def optimize_with_precision_targeting_legacy(shifts_coverage, demand_matrix, *, cfg=None):
    """Optimización de precisión EXACTA del legacy Streamlit."""
    cfg = merge_legacy_config(cfg)
    
    print(f"[LEGACY_PRECISION] Iniciando optimización de precisión")
    
    if not PULP_AVAILABLE:
        return optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, cfg=cfg)
    
    try:
        shifts_list = list(shifts_coverage.keys())
        if not shifts_list:
            return {}, "NO_SHIFTS"
        
        prob = pl.LpProblem("Precision_Scheduling", pl.LpMinimize)
        
        # Variables con límites dinámicos EXACTOS del legacy
        total_demand = demand_matrix.sum()
        peak_demand = demand_matrix.max()
        max_per_shift = max(20, int(total_demand / cfg["agent_limit_factor"]))
        
        shift_vars = {}
        for shift in shifts_list:
            shift_vars[shift] = pl.LpVariable(f"shift_{shift}", 0, max_per_shift, pl.LpInteger)
        
        # Variables de déficit y exceso
        deficit_vars = {}
        excess_vars = {}
        hours = demand_matrix.shape[1]
        patterns_unpacked = {
            s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
            for s, p in shifts_coverage.items()
        }
        
        for day in range(7):
            for hour in range(hours):
                deficit_vars[(day, hour)] = pl.LpVariable(f"deficit_{day}_{hour}", 0, None)
                excess_vars[(day, hour)] = pl.LpVariable(f"excess_{day}_{hour}", 0, None)
        
        # Análisis de patrones críticos EXACTO del legacy
        daily_totals = demand_matrix.sum(axis=1)
        hourly_totals = demand_matrix.sum(axis=0)
        critical_days = np.argsort(daily_totals)[-2:] if len(daily_totals) > 1 else [np.argmax(daily_totals)]
        peak_threshold = np.percentile(hourly_totals[hourly_totals > 0], 75) if np.any(hourly_totals > 0) else 0
        peak_hours = np.where(hourly_totals >= peak_threshold)[0]
        
        # Función objetivo EXACTA del legacy
        total_deficit = pl.lpSum([deficit_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_excess = pl.lpSum([excess_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_agents = pl.lpSum([shift_vars[shift] for shift in shifts_list])
        
        # Bonificaciones por días críticos y horas pico EXACTAS del legacy
        critical_bonus_value = 0
        peak_bonus_value = 0
        
        # Días críticos
        for critical_day in critical_days:
            if critical_day < 7:
                for hour in range(hours):
                    if demand_matrix[critical_day, hour] > 0:
                        critical_bonus_value -= deficit_vars[(critical_day, hour)] * cfg["critical_bonus"]
        
        # Horas pico
        for hour in peak_hours:
            if hour < hours:
                for day in range(7):
                    if demand_matrix[day, hour] > 0:
                        peak_bonus_value -= deficit_vars[(day, hour)] * cfg["peak_bonus"]
        
        # Función objetivo EXACTA del legacy
        prob += (total_deficit * 1000 + 
                 total_excess * cfg["excess_penalty"] + 
                 total_agents * 0.1 + 
                 critical_bonus_value + 
                 peak_bonus_value)
        
        # Restricciones de cobertura EXACTAS del legacy
        for day in range(7):
            for hour in range(hours):
                coverage = pl.lpSum([
                    shift_vars[shift] * patterns_unpacked[shift][day, hour]
                    for shift in shifts_list
                ])
                demand = demand_matrix[day, hour]
                
                prob += coverage + deficit_vars[(day, hour)] >= demand
                prob += coverage - excess_vars[(day, hour)] <= demand
        
        # Restricciones adicionales según perfil EXACTAS del legacy
        if cfg["excess_penalty"] > 5:  # Perfiles estrictos como JEAN
            prob += total_excess <= demand_matrix.sum() * 0.02
        elif cfg["excess_penalty"] > 2:
            prob += total_excess <= demand_matrix.sum() * 0.05
        
        # Límite dinámico de agentes EXACTO del legacy
        dynamic_agent_limit = max(
            int(total_demand / max(1, cfg["agent_limit_factor"])),
            int(peak_demand * 1.1),
        )
        prob += total_agents <= dynamic_agent_limit
        
        # Resolver EXACTO del legacy
        solver = pl.PULP_CBC_CMD(
            msg=cfg.get("solver_msg", 0), 
            timeLimit=cfg.get("solver_time", 240),
            threads=cfg.get("solver_threads", 4)
        )
        prob.solve(solver)
        
        # Extraer solución EXACTA del legacy
        assignments = {}
        if prob.status == pl.LpStatusOptimal:
            for shift in shifts_list:
                value = int(shift_vars[shift].varValue or 0)
                if value > 0:
                    assignments[shift] = value
            method = "PRECISION_TARGETING"
        elif prob.status == pl.LpStatusInfeasible:
            print(f"[LEGACY_PRECISION] Problema infactible, usando greedy")
            return optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, cfg=cfg)
        else:
            print(f"[LEGACY_PRECISION] Solver status: {prob.status}, usando greedy")
            return optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, cfg=cfg)
        
        return assignments, method
        
    except Exception as e:
        print(f"[LEGACY_PRECISION] Error: {e}")
        return optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, cfg=cfg)

def optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, *, cfg=None):
    """Algoritmo greedy mejorado EXACTO del legacy Streamlit."""
    print("[LEGACY_GREEDY] Iniciando algoritmo greedy mejorado")
    cfg = merge_legacy_config(cfg)
    agent_limit_factor = cfg["agent_limit_factor"]
    excess_penalty = cfg["excess_penalty"]
    peak_bonus = cfg["peak_bonus"]
    critical_bonus = cfg["critical_bonus"]

    shifts_list = list(shifts_coverage.keys())
    assignments = {}
    current_coverage = np.zeros_like(demand_matrix, dtype=float)
    
    # Límite de agentes EXACTO del legacy
    max_agents = max(100, int(demand_matrix.sum() / max(1, agent_limit_factor - 5)))
    
    print(f"[LEGACY_GREEDY] Procesando {len(shifts_list)} turnos, max {max_agents} agentes")

    # Análisis de patrones críticos EXACTO del legacy
    daily_totals = demand_matrix.sum(axis=1)
    hourly_totals = demand_matrix.sum(axis=0)
    critical_days = (
        np.argsort(daily_totals)[-2:]
        if daily_totals.size > 1
        else [int(np.argmax(daily_totals))]
    )
    peak_threshold = (
        np.percentile(hourly_totals[hourly_totals > 0], 75)
        if np.any(hourly_totals > 0)
        else 0
    )
    peak_hours = np.where(hourly_totals >= peak_threshold)[0]
    
    print(f"[LEGACY_GREEDY] Días críticos: {critical_days}, Horas pico: {peak_hours}")

    for iteration in range(max_agents):
        if iteration % 20 == 0:
            print(f"[LEGACY_GREEDY] Iteración {iteration}/{max_agents}")
            
        best_shift = None
        best_score = -float("inf")
        best_pattern = None

        for shift_name in shifts_list:
            try:
                # Obtener patrón del turno EXACTO del legacy
                slots_per_day = len(shifts_coverage[shift_name]) // 7
                base_pattern = np.array(shifts_coverage[shift_name]).reshape(7, slots_per_day)
                
                # Ajustar dimensiones si es necesario
                if slots_per_day != demand_matrix.shape[1]:
                    pattern = np.zeros((7, demand_matrix.shape[1]))
                    cols_to_copy = min(slots_per_day, demand_matrix.shape[1])
                    pattern[:, :cols_to_copy] = base_pattern[:, :cols_to_copy]
                else:
                    pattern = base_pattern
                
                new_coverage = current_coverage + pattern
                
                # Cálculo de score EXACTO del legacy
                current_deficit = np.maximum(0, demand_matrix - current_coverage)
                new_deficit = np.maximum(0, demand_matrix - new_coverage)
                deficit_reduction = np.sum(current_deficit - new_deficit)
                
                # Penalización inteligente de exceso EXACTA del legacy
                current_excess = np.maximum(0, current_coverage - demand_matrix)
                new_excess = np.maximum(0, new_coverage - demand_matrix)
                
                smart_excess_penalty = 0
                for day in range(7):
                    for hour in range(demand_matrix.shape[1]):
                        if demand_matrix[day, hour] == 0 and new_excess[day, hour] > current_excess[day, hour]:
                            smart_excess_penalty += 1000  # Penalización extrema
                        elif demand_matrix[day, hour] <= 2:
                            smart_excess_penalty += (new_excess[day, hour] - current_excess[day, hour]) * excess_penalty * 10
                        else:
                            smart_excess_penalty += (new_excess[day, hour] - current_excess[day, hour]) * excess_penalty
                
                # Bonificaciones por patrones críticos EXACTAS del legacy
                critical_bonus_score = 0
                for critical_day in critical_days:
                    if critical_day < 7:
                        day_improvement = np.sum(current_deficit[critical_day] - new_deficit[critical_day])
                        critical_bonus_score += day_improvement * critical_bonus * 2
                
                peak_bonus_score = 0
                for hour in peak_hours:
                    if hour < demand_matrix.shape[1]:
                        hour_improvement = np.sum(current_deficit[:, hour] - new_deficit[:, hour])
                        peak_bonus_score += hour_improvement * peak_bonus * 2
                
                # Score final EXACTO del legacy
                score = (deficit_reduction * 100 + 
                        critical_bonus_score + 
                        peak_bonus_score - 
                        smart_excess_penalty)
                
                if score > best_score:
                    best_score = score
                    best_shift = shift_name
                    best_pattern = pattern
                    
            except Exception as e:
                print(f"[LEGACY_GREEDY] Error procesando {shift_name}: {e}")
                continue
        
        # Criterio de parada EXACTO del legacy
        if best_shift is None or best_score <= 0.5:
            print(f"[LEGACY_GREEDY] Parada en iteración {iteration}, mejor score: {best_score}")
            break
        
        # Aplicar mejor turno
        if best_shift not in assignments:
            assignments[best_shift] = 0
        assignments[best_shift] += 1
        current_coverage += best_pattern
        
        # Verificar si se completó la cobertura
        remaining_deficit = np.sum(np.maximum(0, demand_matrix - current_coverage))
        if remaining_deficit == 0:
            print(f"[LEGACY_GREEDY] Cobertura completa en iteración {iteration}")
            break
    
    total_agents = sum(assignments.values())
    print(f"[LEGACY_GREEDY] Completado: {total_agents} agentes asignados en {len(assignments)} turnos")
    return assignments, "GREEDY_ENHANCED"

def optimize_jean_search_legacy(shifts_coverage, demand_matrix, *, cfg=None, target_coverage=98.0, max_iterations=5):
    """Búsqueda iterativa JEAN EXACTA del legacy original."""
    start_time = time.time()
    max_time = 120  # Máximo 2 minutos para JEAN
    
    cfg = merge_legacy_config(cfg)
    original_factor = cfg["agent_limit_factor"]
    
    best_assignments = {}
    best_method = ""
    best_score = float("inf")
    best_coverage = 0
    
    print(f"[LEGACY_JEAN] Iniciando búsqueda iterativa JEAN")
    
    # Secuencia de factores EXACTA del legacy Streamlit
    factor_sequence = [30, 27, 24, 21, 18, 15, 12, 9, 6, 3]
    factor_sequence = [f for f in factor_sequence if f <= original_factor]
    
    if not factor_sequence:
        factor_sequence = [original_factor]
    
    print(f"[LEGACY_JEAN] Secuencia de factores: {factor_sequence}")
    
    # Implementar EXACTAMENTE la lógica del legacy Streamlit
    for iteration, factor in enumerate(factor_sequence[:max_iterations]):
        # Verificar timeout
        if time.time() - start_time > max_time:
            print(f"[LEGACY_JEAN] Timeout alcanzado ({max_time}s), terminando")
            break
            
        print(f"[LEGACY_JEAN] Iteración {iteration + 1}: factor {factor}")
        
        # Actualizar configuración temporal EXACTA del legacy
        temp_cfg = cfg.copy()
        temp_cfg["agent_limit_factor"] = factor
        
        try:
            assignments, method = optimize_with_precision_targeting_legacy(shifts_coverage, demand_matrix, cfg=temp_cfg)
            results = analyze_results_legacy(assignments, shifts_coverage, demand_matrix)
            
            if results:
                cov = results["coverage_percentage"]
                score = results["overstaffing"] + results["understaffing"]
                print(f"[LEGACY_JEAN] Factor {factor}: cobertura {cov:.1f}%, score {score:.1f}")
                
                if cov >= target_coverage:
                    if score < best_score or not best_assignments:
                        best_assignments, best_method = assignments, f"JEAN_SEARCH_F{factor}"
                        best_score = score
                        best_coverage = cov
                        print(f"[LEGACY_JEAN] Nueva mejor solución: score {score:.1f}")
                    # En JEAN, continuar buscando mejores scores incluso si se alcanza cobertura
                elif cov > best_coverage:
                    best_assignments, best_method, best_coverage = assignments, f"JEAN_SEARCH_F{factor}", cov
                    best_score = score
                    print(f"[LEGACY_JEAN] Mejor cobertura parcial: {cov:.1f}%")
        except Exception as e:
            print(f"[LEGACY_JEAN] Error en iteración {iteration + 1}: {e}")
            # En caso de error, continuar con el siguiente factor
            pass
    
    # Si no hay resultados, usar greedy como fallback
    if not best_assignments:
        print(f"[LEGACY_JEAN] Sin resultados, usando greedy como fallback")
        best_assignments, best_method = optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, cfg=cfg)
    
    elapsed = time.time() - start_time
    print(f"[LEGACY_JEAN] Búsqueda completada en {elapsed:.1f}s: mejor score {best_score:.1f}, cobertura {best_coverage:.1f}%")
    return best_assignments, best_method

def optimize_ft_no_excess_legacy(ft_shifts, demand_matrix, *, cfg=None):
    """Fase 1: FT con CERO exceso permitido EXACTO del legacy."""
    cfg = merge_legacy_config(cfg)
    
    if not ft_shifts:
        return {}
    
    prob = pl.LpProblem("FT_No_Excess", pl.LpMinimize)
    
    # Variables FT
    max_ft_per_shift = max(10, int(demand_matrix.sum() / cfg["agent_limit_factor"]))
    ft_vars = {}
    for shift in ft_shifts.keys():
        ft_vars[shift] = pl.LpVariable(f"ft_{shift}", 0, max_ft_per_shift, pl.LpInteger)
    
    # Solo variables de déficit (NO exceso)
    deficit_vars = {}
    hours = demand_matrix.shape[1]
    patterns_unpacked = {
        s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
        for s, p in ft_shifts.items()
    }
    
    for day in range(7):
        for hour in range(hours):
            deficit_vars[(day, hour)] = pl.LpVariable(f"ft_deficit_{day}_{hour}", 0, None)
    
    # Objetivo: minimizar déficit + agentes
    total_deficit = pl.lpSum([deficit_vars[(day, hour)] for day in range(7) for hour in range(hours)])
    total_ft_agents = pl.lpSum([ft_vars[shift] for shift in ft_shifts.keys()])
    
    prob += total_deficit * 1000 + total_ft_agents * 1
    
    # Restricciones: cobertura <= demanda (SIN exceso)
    for day in range(7):
        for hour in range(hours):
            coverage = pl.lpSum([
                ft_vars[shift] * patterns_unpacked[shift][day, hour]
                for shift in ft_shifts.keys()
            ])
            demand = demand_matrix[day, hour]
            
            # Cobertura + déficit >= demanda
            prob += coverage + deficit_vars[(day, hour)] >= demand
            # Cobertura <= demanda (SIN exceso)
            prob += coverage <= demand
    
    # Resolver
    prob.solve(pl.PULP_CBC_CMD(msg=0, timeLimit=cfg.get("solver_time", 240)//2))
    
    ft_assignments = {}
    if prob.status == pl.LpStatusOptimal:
        for shift in ft_shifts.keys():
            value = int(ft_vars[shift].varValue or 0)
            if value > 0:
                ft_assignments[shift] = value
    
    return ft_assignments

def optimize_pt_complete_legacy(pt_shifts, remaining_demand, *, cfg=None):
    """Fase 2: PT para completar el déficit restante EXACTO del legacy."""
    cfg = merge_legacy_config(cfg)
    
    if not pt_shifts or remaining_demand.sum() == 0:
        return {}
    
    prob = pl.LpProblem("PT_Complete", pl.LpMinimize)
    
    # Variables PT
    max_pt_per_shift = max(10, int(remaining_demand.sum() / max(1, cfg["agent_limit_factor"])))
    pt_vars = {}
    for shift in pt_shifts.keys():
        pt_vars[shift] = pl.LpVariable(f"pt_{shift}", 0, max_pt_per_shift, pl.LpInteger)
    
    # Variables de déficit y exceso
    deficit_vars = {}
    excess_vars = {}
    hours = remaining_demand.shape[1]
    patterns_unpacked = {
        s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
        for s, p in pt_shifts.items()
    }
    
    for day in range(7):
        for hour in range(hours):
            deficit_vars[(day, hour)] = pl.LpVariable(f"pt_deficit_{day}_{hour}", 0, None)
            excess_vars[(day, hour)] = pl.LpVariable(f"pt_excess_{day}_{hour}", 0, None)
    
    # Objetivo: minimizar déficit, controlar exceso
    total_deficit = pl.lpSum([deficit_vars[(day, hour)] for day in range(7) for hour in range(hours)])
    total_excess = pl.lpSum([excess_vars[(day, hour)] for day in range(7) for hour in range(hours)])
    total_pt_agents = pl.lpSum([pt_vars[shift] for shift in pt_shifts.keys()])
    
    prob += total_deficit * 1000 + total_excess * (cfg["excess_penalty"] * 20) + total_pt_agents * 1

    # Para el perfil JEAN no se permite ningún exceso
    prob += total_excess == 0
    
    # Restricciones de cobertura
    for day in range(7):
        for hour in range(hours):
            coverage = pl.lpSum([
                pt_vars[shift] * patterns_unpacked[shift][day, hour]
                for shift in pt_shifts.keys()
            ])
            demand = remaining_demand[day, hour]
            
            prob += coverage + deficit_vars[(day, hour)] >= demand
            prob += coverage - excess_vars[(day, hour)] <= demand
    
    # Resolver
    prob.solve(pl.PULP_CBC_CMD(msg=0, timeLimit=cfg.get("solver_time", 240)//2))
    
    pt_assignments = {}
    if prob.status == pl.LpStatusOptimal:
        for shift in pt_shifts.keys():
            value = int(pt_vars[shift].varValue or 0)
            if value > 0:
                pt_assignments[shift] = value
    
    return pt_assignments

def optimize_ft_then_pt_strategy_legacy(shifts_coverage, demand_matrix, *, cfg=None):
    """Estrategia 2 fases EXACTA del legacy: FT sin exceso, luego PT para completar."""
    print("[LEGACY_FT_PT] Iniciando estrategia 2 fases")
    cfg = merge_legacy_config(cfg)
    
    ft_shifts = {k: v for k, v in shifts_coverage.items() if k.startswith('FT')}
    pt_shifts = {k: v for k, v in shifts_coverage.items() if k.startswith('PT')}
    
    print(f"[LEGACY_FT_PT] Fase 1: {len(ft_shifts)} turnos FT")
    ft_assignments = optimize_ft_no_excess_legacy(ft_shifts, demand_matrix, cfg=cfg)
    
    # Calcular cobertura FT
    ft_coverage = np.zeros_like(demand_matrix)
    for name, count in ft_assignments.items():
        pattern = ft_shifts[name].reshape(7, demand_matrix.shape[1]) if len(ft_shifts[name]) == 7 * demand_matrix.shape[1] else ft_shifts[name].reshape(7, -1)[:, :demand_matrix.shape[1]]
        ft_coverage += pattern * count
    
    print(f"[LEGACY_FT_PT] Fase 2: {len(pt_shifts)} turnos PT")
    remaining_demand = np.maximum(0, demand_matrix - ft_coverage)
    pt_assignments = optimize_pt_complete_legacy(pt_shifts, remaining_demand, cfg=cfg)
    
    final_assignments = {**ft_assignments, **pt_assignments}
    print(f"[LEGACY_FT_PT] Completado: {len(final_assignments)} turnos asignados")
    return final_assignments, "FT_NO_EXCESS_THEN_PT"

def optimize_schedule_iterative_legacy(shifts_coverage, demand_matrix, *, cfg=None):
    """Función principal EXACTA del legacy Streamlit con estrategia FT primero + PT después."""
    cfg = merge_legacy_config(cfg)
    
    if PULP_AVAILABLE:
        profile = cfg.get("optimization_profile", "JEAN")
        
        if profile == "JEAN":
            print("[LEGACY] Búsqueda JEAN: cobertura sin exceso")
            return optimize_jean_search_legacy(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=cfg.get("TARGET_COVERAGE", 98.0))

        if profile == "JEAN Personalizado":
            if cfg.get("use_ft") and cfg.get("use_pt"):
                print("[LEGACY] Estrategia 2 Fases: FT sin exceso -> PT para completar")
                assignments, method = optimize_ft_then_pt_strategy_legacy(shifts_coverage, demand_matrix, cfg=cfg)

                results = analyze_results_legacy(assignments, shifts_coverage, demand_matrix)
                if results:
                    cov = results["coverage_percentage"]
                    score = results["overstaffing"] + results["understaffing"]
                    target_coverage = cfg.get("TARGET_COVERAGE", 98.0)
                    if cov < target_coverage or score > 0:
                        print("[LEGACY] Refinando con búsqueda JEAN")
                        assignments, method = optimize_jean_search_legacy(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=target_coverage)

                return assignments, method
            else:
                print("[LEGACY] Búsqueda JEAN: cobertura sin exceso")
                return optimize_jean_search_legacy(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=cfg.get("TARGET_COVERAGE", 98.0))

        if cfg.get("use_ft") and cfg.get("use_pt"):
            print("[LEGACY] Estrategia 2 Fases: FT sin exceso -> PT para completar")
            return optimize_ft_then_pt_strategy_legacy(shifts_coverage, demand_matrix, cfg=cfg)
        else:
            print("[LEGACY] Modo Precisión: Optimización directa")
            return optimize_with_precision_targeting_legacy(shifts_coverage, demand_matrix, cfg=cfg)
    else:
        print("[LEGACY] Solver Básico: Greedy mejorado")
        return optimize_schedule_greedy_enhanced_legacy(shifts_coverage, demand_matrix, cfg=cfg)

def solve_in_chunks_optimized_legacy(shifts_coverage, demand_matrix, *, cfg=None):
    """Solve EXACTO del legacy Streamlit usando chunks ordenados por score."""
    scored = []
    seen = set()
    for name, pat in shifts_coverage.items():
        key = hashlib.md5(pat).digest()
        if key in seen:
            continue
        seen.add(key)
        scored.append((name, pat, score_pattern(pat, demand_matrix)))

    scored.sort(key=lambda x: x[2], reverse=True)

    assignments_total = {}
    coverage = np.zeros_like(demand_matrix)
    idx = 0
    base_chunk_size = 10000
    
    while idx < len(scored):
        chunk_size = adaptive_chunk_size(base_chunk_size)
        chunk_dict = {name: pat for name, pat, _ in scored[idx: idx + chunk_size]}
        remaining = np.maximum(0, demand_matrix - coverage)
        if not np.any(remaining):
            break
        
        # Usar la función de optimización según el perfil
        assigns, _ = optimize_schedule_iterative_legacy(chunk_dict, remaining, cfg=cfg)
        
        for name, val in assigns.items():
            assignments_total[name] = assignments_total.get(name, 0) + val
            slots = len(chunk_dict[name]) // 7
            pat_matrix = chunk_dict[name].reshape(7, slots)
            if slots != demand_matrix.shape[1]:
                # Ajustar dimensiones si es necesario
                adjusted_pattern = np.zeros((7, demand_matrix.shape[1]))
                cols_to_copy = min(slots, demand_matrix.shape[1])
                adjusted_pattern[:, :cols_to_copy] = pat_matrix[:, :cols_to_copy]
                pat_matrix = adjusted_pattern
            coverage += pat_matrix * val
        
        idx += chunk_size
        gc.collect()
        emergency_cleanup()
        
        if not np.any(np.maximum(0, demand_matrix - coverage)):
            break
    
    return assignments_total

def analyze_results_legacy(assignments, shifts_coverage, demand_matrix, coverage_matrix=None):
    """Compute coverage metrics from solved assignments - EXACTO del legacy."""
    if not assignments:
        return None

    compute_coverage = coverage_matrix is None
    if compute_coverage:
        # Usar las mismas dimensiones que demand_matrix
        coverage_matrix = np.zeros_like(demand_matrix, dtype=np.int16)
    
    total_agents = 0
    ft_agents = 0
    pt_agents = 0
    
    for shift_name, count in assignments.items():
        total_agents += count
        if shift_name.startswith('FT'):
            ft_agents += count
        else:
            pt_agents += count
            
        if compute_coverage and shift_name in shifts_coverage:
            weekly_pattern = shifts_coverage[shift_name]
            target_shape = demand_matrix.shape
            
            # Conversión de dimensiones EXACTA del legacy
            if len(weekly_pattern) == target_shape[0] * target_shape[1]:
                pattern_matrix = weekly_pattern.reshape(target_shape)
            else:
                # Reshape temporal y ajuste
                slots_per_day = len(weekly_pattern) // 7
                pattern_temp = weekly_pattern.reshape(7, slots_per_day)
                pattern_matrix = np.zeros(target_shape)
                
                if slots_per_day == target_shape[1]:
                    # Dimensiones coinciden
                    pattern_matrix = pattern_temp
                else:
                    # Ajustar dimensiones copiando lo que se pueda
                    cols_to_copy = min(slots_per_day, target_shape[1])
                    pattern_matrix[:, :cols_to_copy] = pattern_temp[:, :cols_to_copy]
            
            coverage_matrix += pattern_matrix * count

    # Cálculo de métricas EXACTO del legacy
    total_demand = demand_matrix.sum()
    total_covered = np.minimum(coverage_matrix, demand_matrix).sum()
    
    # Coverage percentage: demanda cubierta / demanda total (ponderado por horas)
    coverage_percentage = (
        (total_covered / total_demand) * 100 if total_demand > 0 else 0
    )
    
    diff_matrix = coverage_matrix - demand_matrix
    overstaffing = np.sum(diff_matrix[diff_matrix > 0])
    understaffing = np.sum(np.abs(diff_matrix[diff_matrix < 0]))
    
    return {
        'total_coverage': coverage_matrix,
        'total_agents': total_agents,
        'ft_agents': ft_agents,
        'pt_agents': pt_agents,
        'coverage_percentage': coverage_percentage,
        'overstaffing': overstaffing,
        'understaffing': understaffing,
        'diff_matrix': diff_matrix,
    }

def create_heatmap_legacy(matrix, title, cmap="RdYlBu_r"):
    """Return a matplotlib figure with matrix visualised as a heatmap."""
    if plt is None:
        return None
        
    fig, ax = plt.subplots(figsize=(12, 6))
    im = ax.imshow(matrix, cmap=cmap, aspect="auto")
    ax.set_xticks(range(24))
    ax.set_xticklabels([f"{h:02d}" for h in range(24)])
    ax.set_yticks(range(7))
    ax.set_yticklabels([
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    ])
    for i in range(7):
        for j in range(24):
            ax.text(j, i, f"{matrix[i, j]:.0f}", ha="center", va="center", color="black", fontsize=8)
    ax.set_title(title)
    ax.set_xlabel("Hora del día")
    ax.set_ylabel("Día de la semana")
    plt.colorbar(im, ax=ax)
    return fig

def generate_all_heatmaps_legacy(demand, coverage=None, diff=None) -> dict:
    """Generate heatmaps for demand, coverage and difference matrices."""
    maps = {"demand": create_heatmap_legacy(demand, "Demanda por Hora y Día", "Reds")}
    if coverage is not None:
        maps["coverage"] = create_heatmap_legacy(coverage, "Cobertura por Hora y Día", "Blues")
    if diff is not None:
        maps["difference"] = create_heatmap_legacy(diff, "Diferencias por Hora y Día", "RdBu")
    return maps

def _extract_start_hour_legacy(name: str) -> float:
    """Best-effort extraction of the start hour from a shift name."""
    for part in name.split('_'):
        if '.' in part and part.replace('.', '').isdigit():
            try:
                return float(part)
            except ValueError:
                continue
    return 0.0

def export_detailed_schedule_legacy(assignments, shifts_coverage):
    """Return schedule data in Excel and CSV formats - EXACTO del legacy."""
    if not assignments:
        return None, None

    try:
        from openpyxl import Workbook
        from collections import defaultdict
    except ImportError:
        return None, None

    DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

    wb = Workbook(write_only=True)
    ws_details = wb.create_sheet('Horarios_Semanales')
    ws_summary = wb.create_sheet('Resumen_Agentes')
    ws_shifts = wb.create_sheet('Turnos_Asignados')

    ws_details.append(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])
    ws_summary.append(['Agente', 'Turno', 'Dias_Trabajo'])
    ws_shifts.append(['Turno', 'Agentes'])

    csv_buffer = StringIO()
    csv_writer = csv.writer(csv_buffer)
    csv_writer.writerow(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])

    summary_counts = defaultdict(int)

    agent_id = 1
    for shift_name, count in assignments.items():
        weekly_pattern = shifts_coverage[shift_name]
        slots_per_day = 24  # Always use 24 hours per day
        pattern_matrix = weekly_pattern.reshape(7, slots_per_day) if len(weekly_pattern) == 7 * slots_per_day else weekly_pattern.reshape(7, -1)[:, :slots_per_day]
        parts = shift_name.split('_')
        start_hour = _extract_start_hour_legacy(shift_name)
        
        if shift_name.startswith('FT10p8'):
            shift_type = 'FT'
            shift_duration = 10
            total_hours = shift_duration + 1
        elif shift_name.startswith('FT'):
            shift_type = 'FT'
            if '_Variado_' in shift_name or len(parts) > 4:
                shift_duration = int(pattern_matrix.sum(axis=1).max())
            else:
                try:
                    if len(parts[0]) > 2 and parts[0][2:].isdigit():
                        shift_duration = int(parts[0][2:])
                    else:
                        shift_duration = int(pattern_matrix.sum(axis=1).max())
                except (ValueError, IndexError):
                    shift_duration = int(pattern_matrix.sum(axis=1).max())
            total_hours = shift_duration + 1
        elif shift_name.startswith('PT'):
            shift_type = 'PT'
            if '_Variado_' in shift_name or len(parts) > 4:
                shift_duration = int(pattern_matrix.sum(axis=1).max())
            else:
                try:
                    if len(parts[0]) > 2 and parts[0][2:].isdigit():
                        shift_duration = int(parts[0][2:])
                    else:
                        shift_duration = int(pattern_matrix.sum(axis=1).max())
                except (ValueError, IndexError):
                    shift_duration = int(pattern_matrix.sum(axis=1).max())
            total_hours = shift_duration
        else:
            shift_type = 'FT'
            shift_duration = 8
            total_hours = 9

        for _ in range(count):
            for day in range(7):
                day_pattern = pattern_matrix[day]
                work_hours = np.where(day_pattern == 1)[0]
                if len(work_hours) > 0:
                    start_idx = int(start_hour)
                    if shift_name.startswith('PT') or shift_name.startswith('FT10p8'):
                        end_idx = (int(work_hours[-1]) + 1) % 24
                        next_day = end_idx <= start_idx
                        horario = f"{start_idx:02d}:00-{end_idx:02d}:00" + ("+1" if next_day else "")
                    else:
                        end_idx = int(work_hours[-1]) + 1
                        next_day = end_idx <= start_idx
                        horario = f"{start_idx:02d}:00-{end_idx % 24:02d}:00" + ("+1" if next_day else "")
                    
                    if shift_name.startswith('PT'):
                        break_time = ""
                    elif shift_name.startswith('FT10p8'):
                        all_expected = set(range(int(start_hour), int(start_hour + total_hours)))
                        actual_hours = set(work_hours)
                        break_hours = all_expected - actual_hours
                        if break_hours:
                            break_hour = min(break_hours) % 24
                            break_end = (break_hour + 1) % 24 or 24
                            break_time = f"{break_hour:02d}:00-{break_end:02d}:00"
                        else:
                            break_time = ""
                    else:
                        expected = list(range(start_idx, end_idx))
                        if next_day:
                            expected = list(range(start_idx, 24)) + list(range(0, end_idx % 24))
                        break_hours = set(expected) - set(work_hours)
                        if break_hours:
                            break_hour = min(break_hours)
                            break_time = f"{break_hour % 24:02d}:00-{((break_hour + 1) % 24) or 24:02d}:00"
                        else:
                            break_time = ""
                    
                    row = [f"AGT_{agent_id:03d}", DAYS[day], horario, break_time, shift_name, shift_type]
                else:
                    row = [f"AGT_{agent_id:03d}", DAYS[day], "DSO", "", shift_name, "DSO"]
                
                ws_details.append(row)
                csv_writer.writerow(row)
                summary_counts[(agent_id, shift_name)] += 1
            agent_id += 1

    for (agent_idx, shift_name), cnt in summary_counts.items():
        ws_summary.append([f"AGT_{agent_idx:03d}", shift_name, cnt])

    for shift, count in assignments.items():
        ws_shifts.append([shift, count])

    excel_io = BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()

    csv_bytes = csv_buffer.getvalue().encode("utf-8")
    return excel_bytes, csv_bytes

def run_legacy_optimization(demand_matrix, config=None, generate_charts=False):
    """Función principal EXACTA del legacy Streamlit."""
    print("[LEGACY_MAIN] Iniciando optimización legacy")
    
    cfg = merge_legacy_config(config)
    
    # Generar patrones EXACTO del legacy
    print("[LEGACY_MAIN] Generando patrones de turnos...")
    shifts_coverage = generate_shifts_coverage_legacy(cfg=cfg)
    
    if not shifts_coverage:
        print("[LEGACY_MAIN] ERROR: No se pudieron generar patrones")
        return {"error": "No se pudieron generar patrones válidos"}, None, None
    
    print(f"[LEGACY_MAIN] Generados {len(shifts_coverage)} patrones")
    
    # Optimizar usando chunks EXACTO del legacy
    print("[LEGACY_MAIN] Optimizando con chunks...")
    assignments = solve_in_chunks_optimized_legacy(shifts_coverage, demand_matrix, cfg=cfg)
    
    if not assignments:
        print("[LEGACY_MAIN] ERROR: No se pudo encontrar solución")
        return {"error": "No se pudo encontrar una solución válida"}, None, None
    
    # Analizar resultados EXACTO del legacy
    print("[LEGACY_MAIN] Analizando resultados...")
    metrics = analyze_results_legacy(assignments, shifts_coverage, demand_matrix)
    
    # Exportar archivos EXACTO del legacy
    excel_bytes, csv_bytes = export_detailed_schedule_legacy(assignments, shifts_coverage)
    
    # Generar gráficas si se solicita
    heatmaps = {}
    if generate_charts and metrics:
        maps = generate_all_heatmaps_legacy(
            demand_matrix,
            metrics.get("total_coverage"),
            metrics.get("diff_matrix")
        )
        for key, fig in maps.items():
            if fig:
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                fig.savefig(tmp.name, format="png", bbox_inches="tight")
                tmp.flush()
                tmp.close()
                filename = tmp.name.split('/')[-1] if '/' in tmp.name else tmp.name.split('\\')[-1]
                heatmaps[key] = f"/heatmap/{filename}"
                plt.close(fig)
    
    # Preparar resultado EXACTO del legacy
    result = {
        "assignments": assignments,
        "metrics": metrics,
        "heatmaps": heatmaps,
        "status": "LEGACY_OPTIMIZED",
        "pulp_results": {
            "assignments": assignments,
            "metrics": metrics,
            "status": "LEGACY_PULP",
            "heatmaps": heatmaps
        },
        "greedy_results": {
            "assignments": {},
            "metrics": None,
            "status": "NOT_EXECUTED",
            "heatmaps": {}
        }
    }
    
    print(f"[LEGACY_MAIN] Optimización completada: {sum(assignments.values())} agentes en {len(assignments)} turnos")
    
    return result, excel_bytes, csv_bytes