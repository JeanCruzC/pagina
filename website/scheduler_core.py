# -*- coding: utf-8 -*-
"""
Scheduler Core - Lógica EXACTA 1:1 del generador Streamlit original
"""
import json
import time
import os
import gc
import hashlib
import numpy as np
import pandas as pd
from io import BytesIO, StringIO
from itertools import combinations, permutations
import tempfile
import csv
from collections import defaultdict

try:
    import pulp as pl
    PULP_AVAILABLE = True
except Exception:
    PULP_AVAILABLE = False

try:
    import psutil
except Exception:
    psutil = None

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns
except Exception:
    matplotlib = None
    plt = None
    sns = None

print(f"[CORE] PuLP disponible: {PULP_AVAILABLE}")

# Configuración por defecto EXACTA del Streamlit original
DEFAULT_CONFIG = {
    "solver_time": 300,
    "solver_msg": 1,
    "TARGET_COVERAGE": 98.0,
    "agent_limit_factor": 12,
    "excess_penalty": 2.0,
    "peak_bonus": 1.5,
    "critical_bonus": 2.0,
    "iterations": 30,
    "solver_threads": os.cpu_count() or 1,
    "use_ft": True,
    "use_pt": True,
    "allow_8h": True,
    "allow_10h8": False,
    "allow_pt_4h": True,
    "allow_pt_6h": True,
    "allow_pt_5h": False,
    "break_from_start": 2.5,
    "break_from_end": 2.5,
    "ACTIVE_DAYS": list(range(7)),
}

def merge_config(cfg=None):
    """Combinar configuración por defecto con la proporcionada."""
    merged = DEFAULT_CONFIG.copy()
    if cfg:
        merged.update({k: v for k, v in cfg.items() if v is not None})
    return merged

def monitor_memory_usage():
    """Retornar porcentaje de uso de memoria."""
    if psutil is None:
        return 0.0
    return psutil.virtual_memory().percent

def load_demand_matrix_from_df(df) -> np.ndarray:
    """Cargar matriz de demanda desde DataFrame - EXACTO del original."""
    demand_matrix = np.zeros((7, 24), dtype=float)
    
    day_col = "Día"
    time_col = "Horario"
    demand_col = "Suma de Agentes Requeridos Erlang"
    
    # Buscar columnas si no existen con nombres exactos
    if day_col not in df.columns or time_col not in df.columns or demand_col not in df.columns:
        for col in df.columns:
            if "día" in col.lower() or "dia" in col.lower():
                day_col = col
            elif "horario" in col.lower():
                time_col = col
            elif "erlang" in col.lower() or "requeridos" in col.lower():
                demand_col = col
    
    for _, row in df.iterrows():
        try:
            day = int(row[day_col])
            if not (1 <= day <= 7):
                continue
            day_idx = day - 1
            
            horario = str(row[time_col])
            if ":" in horario:
                hour = int(horario.split(":")[0])
            else:
                hour = int(float(horario))
            if not (0 <= hour <= 23):
                continue
            
            demanda = float(row[demand_col])
            demand_matrix[day_idx, hour] = demanda
        except (ValueError, TypeError, IndexError):
            continue
    
    return demand_matrix

def analyze_demand_matrix(matrix: np.ndarray) -> dict:
    """Analizar matriz de demanda - EXACTO del original."""
    daily_demand = matrix.sum(axis=1)
    hourly_demand = matrix.sum(axis=0)
    active_days = [d for d in range(7) if daily_demand[d] > 0]
    inactive_days = [d for d in range(7) if daily_demand[d] == 0]
    working_days = len(active_days)
    active_hours = np.where(hourly_demand > 0)[0]
    first_hour = int(active_hours.min()) if active_hours.size else 8
    last_hour = int(active_hours.max()) if active_hours.size else 20
    operating_hours = last_hour - first_hour + 1
    peak_demand = float(matrix.max()) if matrix.size else 0.0
    avg_demand = float(matrix[active_days].mean()) if active_days else 0.0
    
    daily_totals = matrix.sum(axis=1)
    hourly_totals = matrix.sum(axis=0)
    critical_days = (
        np.argsort(daily_totals)[-2:] if daily_totals.size > 1 else [int(np.argmax(daily_totals))]
    )
    peak_threshold = (
        np.percentile(hourly_totals[hourly_totals > 0], 75)
        if np.any(hourly_totals > 0)
        else 0
    )
    peak_hours = np.where(hourly_totals >= peak_threshold)[0]
    
    return {
        "daily_demand": daily_demand,
        "hourly_demand": hourly_demand,
        "active_days": active_days,
        "inactive_days": inactive_days,
        "working_days": working_days,
        "first_hour": first_hour,
        "last_hour": last_hour,
        "operating_hours": operating_hours,
        "peak_demand": peak_demand,
        "average_demand": avg_demand,
        "critical_days": critical_days,
        "peak_hours": peak_hours,
    }

def generate_weekly_pattern(start_hour, duration, working_days, dso_day=None, break_len=1, *, cfg=None):
    """Generar patrón semanal - EXACTO del original."""
    cfg = merge_config(cfg)
    break_from_start = cfg["break_from_start"]
    break_from_end = cfg["break_from_end"]
    pattern = np.zeros((7, 24), dtype=np.int8)
    
    for day in working_days:
        if day != dso_day:
            for h in range(duration):
                t = start_hour + h
                d_off, idx = divmod(int(t), 24)
                pattern[(day + d_off) % 7, idx] = 1
            
            # Calcular break
            break_start_idx = start_hour + break_from_start
            break_end_idx = start_hour + duration - break_from_end
            if int(break_start_idx) < int(break_end_idx):
                break_hour = int(break_start_idx) + (int(break_end_idx) - int(break_start_idx)) // 2
            else:
                break_hour = int(break_start_idx)
            
            for b in range(int(break_len)):
                t = break_hour + b
                d_off, idx = divmod(int(t), 24)
                pattern[(day + d_off) % 7, idx] = 0
    
    return pattern.flatten()

def generate_weekly_pattern_simple(start_hour, duration, working_days):
    """Patrón simple sin breaks - EXACTO del original."""
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        for h in range(duration):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
    return pattern.flatten()

def generate_shifts_coverage(*, cfg=None):
    """Generar todos los patrones de turnos - EXACTO del original."""
    cfg = merge_config(cfg)
    use_ft = cfg["use_ft"]
    use_pt = cfg["use_pt"]
    allow_8h = cfg["allow_8h"]
    allow_10h8 = cfg["allow_10h8"]
    allow_pt_4h = cfg["allow_pt_4h"]
    allow_pt_6h = cfg["allow_pt_6h"]
    allow_pt_5h = cfg["allow_pt_5h"]
    active_days = cfg["ACTIVE_DAYS"]
    
    shifts_coverage = {}
    seen_patterns = set()
    step = 0.5
    start_hours = [h for h in np.arange(0, 24, step) if h <= 23.5]
    
    for start_hour in start_hours:
        # Full Time 8 horas - 6 días de trabajo
        if use_ft and allow_8h:
            for dso_day in active_days:
                working_days = [d for d in active_days if d != dso_day][:6]
                if len(working_days) >= 6 and 8 * len(working_days) <= 48:
                    pattern = generate_weekly_pattern(start_hour, 8, working_days, dso_day, cfg=cfg)
                    key = pattern.tobytes()
                    if key not in seen_patterns:
                        seen_patterns.add(key)
                        name = f"FT8_{start_hour:04.1f}_DSO{dso_day}"
                        shifts_coverage[name] = pattern
        
        # Part Time patterns
        if use_pt:
            # 4 horas - múltiples combinaciones de días
            if allow_pt_4h:
                for num_days in [4, 5, 6]:
                    if num_days <= len(active_days) and 4 * num_days <= 24:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 4, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT4_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
            
            # 6 horas - 4 días (24h/semana)
            if allow_pt_6h:
                for num_days in [4]:
                    if num_days <= len(active_days) and 6 * num_days <= 24:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 6, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT6_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
    
    return shifts_coverage

def analyze_results(assignments, shifts_coverage, demand_matrix):
    """Analizar resultados - EXACTO del original."""
    if not assignments:
        return None
    
    coverage_matrix = np.zeros_like(demand_matrix, dtype=np.int16)
    total_agents = 0
    ft_agents = 0
    pt_agents = 0
    
    for shift_name, count in assignments.items():
        total_agents += count
        if shift_name.startswith('FT'):
            ft_agents += count
        else:
            pt_agents += count
        
        if shift_name in shifts_coverage:
            weekly_pattern = shifts_coverage[shift_name]
            target_shape = demand_matrix.shape
            
            if len(weekly_pattern) == target_shape[0] * target_shape[1]:
                pattern_matrix = weekly_pattern.reshape(target_shape)
            else:
                slots_per_day = len(weekly_pattern) // 7
                pattern_temp = weekly_pattern.reshape(7, slots_per_day)
                pattern_matrix = np.zeros(target_shape)
                
                if slots_per_day == target_shape[1]:
                    pattern_matrix = pattern_temp
                else:
                    cols_to_copy = min(slots_per_day, target_shape[1])
                    pattern_matrix[:, :cols_to_copy] = pattern_temp[:, :cols_to_copy]
            
            coverage_matrix += pattern_matrix * count
    
    # Cálculo de métricas EXACTO del original
    total_demand = demand_matrix.sum()
    total_covered = np.minimum(coverage_matrix, demand_matrix).sum()
    
    coverage_percentage = (
        (total_covered / total_demand) * 100 if total_demand > 0 else 0
    )
    
    diff_matrix = coverage_matrix - demand_matrix
    overstaffing = np.sum(diff_matrix[diff_matrix > 0])
    understaffing = np.sum(np.abs(diff_matrix[diff_matrix < 0]))
    
    return {
        'total_coverage': coverage_matrix,
        'total_agents': total_agents,
        'ft_agents': ft_agents,
        'pt_agents': pt_agents,
        'coverage_percentage': coverage_percentage,
        'overstaffing': overstaffing,
        'understaffing': understaffing,
        'diff_matrix': diff_matrix,
    }

def create_heatmap(matrix, title, cmap="RdYlBu_r"):
    """Crear heatmap - EXACTO del original."""
    if plt is None:
        return None
    
    fig, ax = plt.subplots(figsize=(12, 6))
    im = ax.imshow(matrix, cmap=cmap, aspect="auto")
    ax.set_xticks(range(24))
    ax.set_xticklabels([f"{h:02d}" for h in range(24)])
    ax.set_yticks(range(7))
    ax.set_yticklabels([
        "Lunes", "Martes", "Miércoles", "Jueves", 
        "Viernes", "Sábado", "Domingo"
    ])
    
    for i in range(7):
        for j in range(24):
            ax.text(j, i, f"{matrix[i, j]:.0f}", ha="center", va="center", 
                   color="black", fontsize=8)
    
    ax.set_title(title)
    ax.set_xlabel("Hora del día")
    ax.set_ylabel("Día de la semana")
    plt.colorbar(im, ax=ax)
    return fig

def generate_all_heatmaps(demand, coverage=None, diff=None) -> dict:
    """Generar todos los heatmaps."""
    maps = {}
    if plt is not None:
        maps["demand"] = create_heatmap(demand, "Demanda por Hora y Día", "Reds")
        if coverage is not None:
            maps["coverage"] = create_heatmap(coverage, "Cobertura por Hora y Día", "Blues")
        if diff is not None:
            maps["difference"] = create_heatmap(diff, "Diferencias por Hora y Día", "RdBu")
    return maps

def export_detailed_schedule(assignments, shifts_coverage):
    """Exportar horarios detallados - EXACTO del original."""
    if not assignments:
        return None, None
    
    try:
        from openpyxl import Workbook
    except ImportError:
        return None, None
    
    DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    
    wb = Workbook(write_only=True)
    ws_details = wb.create_sheet('Horarios_Semanales')
    ws_summary = wb.create_sheet('Resumen_Agentes')
    ws_shifts = wb.create_sheet('Turnos_Asignados')
    
    ws_details.append(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])
    ws_summary.append(['Agente', 'Turno', 'Dias_Trabajo'])
    ws_shifts.append(['Turno', 'Agentes'])
    
    csv_buffer = StringIO()
    csv_writer = csv.writer(csv_buffer)
    csv_writer.writerow(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])
    
    summary_counts = defaultdict(int)
    agent_id = 1
    
    for shift_name, count in assignments.items():
        weekly_pattern = shifts_coverage[shift_name]
        slots_per_day = 24
        pattern_matrix = weekly_pattern.reshape(7, slots_per_day) if len(weekly_pattern) == 7 * slots_per_day else weekly_pattern.reshape(7, -1)[:, :slots_per_day]
        
        # Extraer hora de inicio del nombre del turno
        start_hour = 0.0
        for part in shift_name.split('_'):
            if '.' in part and part.replace('.', '').isdigit():
                try:
                    start_hour = float(part)
                    break
                except ValueError:
                    continue
        
        shift_type = 'FT' if shift_name.startswith('FT') else 'PT'
        
        for _ in range(count):
            for day in range(7):
                day_pattern = pattern_matrix[day]
                work_hours = np.where(day_pattern == 1)[0]
                
                if len(work_hours) > 0:
                    start_idx = int(start_hour)
                    end_idx = int(work_hours[-1]) + 1
                    next_day = end_idx <= start_idx
                    horario = f"{start_idx:02d}:00-{end_idx % 24:02d}:00" + ("+1" if next_day else "")
                    
                    # Calcular break para FT
                    if shift_name.startswith('FT'):
                        expected = list(range(start_idx, end_idx))
                        if next_day:
                            expected = list(range(start_idx, 24)) + list(range(0, end_idx % 24))
                        break_hours = set(expected) - set(work_hours)
                        if break_hours:
                            break_hour = min(break_hours)
                            break_time = f"{break_hour % 24:02d}:00-{((break_hour + 1) % 24) or 24:02d}:00"
                        else:
                            break_time = ""
                    else:
                        break_time = ""
                    
                    row = [f"AGT_{agent_id:03d}", DAYS[day], horario, break_time, shift_name, shift_type]
                else:
                    row = [f"AGT_{agent_id:03d}", DAYS[day], "DSO", "", shift_name, "DSO"]
                
                ws_details.append(row)
                csv_writer.writerow(row)
                summary_counts[(agent_id, shift_name)] += 1
            agent_id += 1
    
    for (agent_idx, shift_name), cnt in summary_counts.items():
        ws_summary.append([f"AGT_{agent_idx:03d}", shift_name, cnt])
    
    for shift, count in assignments.items():
        ws_shifts.append([shift, count])
    
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()
    
    csv_bytes = csv_buffer.getvalue().encode("utf-8")
    return excel_bytes, csv_bytes