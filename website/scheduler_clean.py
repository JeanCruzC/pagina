# -*- coding: utf-8 -*-
"""
Scheduler Clean - Fixed version with all security and performance issues resolved
"""
import json
import time
import os
import gc
import hashlib
import numpy as np
import pandas as pd
from io import BytesIO
from itertools import combinations, permutations
import tempfile
from collections import defaultdict

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns
except ImportError:
    matplotlib = None
    plt = None
    sns = None

try:
    import psutil
except ImportError:
    psutil = None

try:
    import pulp
    PULP_AVAILABLE = True
except ImportError:
    PULP_AVAILABLE = False

from threading import RLock
from functools import wraps

_MODEL_LOCK = RLock()
active_jobs = {}

# Constants
MAX_FT_HOURS = 48
MAX_PT_HOURS = 24

def monitor_memory_usage() -> float:
    """Return current memory usage percentage."""
    if psutil is None:
        return 0.0
    return psutil.virtual_memory().percent

def memory_limit_patterns(slots_per_day: int) -> int:
    """Return how many patterns fit in roughly 4GB of RAM."""
    if slots_per_day <= 0:
        return 0
    if psutil is None:
        return 10000
    available = psutil.virtual_memory().available
    cap = min(available, 4 * 1024 ** 3)
    return int(cap // (7 * slots_per_day))

def adaptive_chunk_size(base: int = 5000) -> int:
    """Return dynamic chunk size based on memory pressure."""
    if psutil is None:
        return base
    usage = monitor_memory_usage()
    if usage > 80:
        return max(1000, base // 4)
    if usage > 60:
        return max(2000, base // 2)
    return base

def emergency_cleanup(threshold: float = 85.0) -> bool:
    """Trigger garbage collection if memory usage exceeds threshold."""
    if psutil is None:
        return False
    if monitor_memory_usage() >= threshold:
        gc.collect()
        return True
    return False

def create_demand_signature(demand_matrix):
    """Create unique signature for demand pattern."""
    normalized = demand_matrix / (demand_matrix.max() + 1e-8)
    signature = hashlib.sha256(normalized.tobytes()).hexdigest()[:16]
    return signature

def get_adaptive_params(demand_matrix, target_coverage):
    """Get adaptive parameters based on demand pattern."""
    input_hash = hashlib.sha256(str(demand_matrix).encode()).hexdigest()[:12]
    
    total_demand = demand_matrix.sum()
    peak_demand = demand_matrix.max()
    
    return {
        "agent_limit_factor": max(8, int(total_demand / max(1, peak_demand) * 3)),
        "excess_penalty": 0.05,
        "peak_bonus": 2.5,
        "critical_bonus": 3.0,
        "precision_mode": True,
        "learned": False,
        "runs_count": 0,
        "evolution_step": "initial"
    }

def analyze_demand_matrix(matrix: np.ndarray) -> dict:
    """Analyze demand matrix and return metrics."""
    daily_demand = matrix.sum(axis=1)
    hourly_demand = matrix.sum(axis=0)
    active_days = [d for d in range(7) if daily_demand[d] > 0]
    inactive_days = [d for d in range(7) if daily_demand[d] == 0]
    working_days = len(active_days)
    active_hours = np.where(hourly_demand > 0)[0]
    first_hour = int(active_hours.min()) if active_hours.size else 8
    last_hour = int(active_hours.max()) if active_hours.size else 20
    operating_hours = last_hour - first_hour + 1
    peak_demand = float(matrix.max()) if matrix.size else 0.0
    avg_demand = float(matrix[active_days].mean()) if active_days else 0.0
    
    daily_totals = matrix.sum(axis=1)
    hourly_totals = matrix.sum(axis=0)
    
    if daily_totals.size == 0 or daily_totals.max() == 0:
        critical_days = []
    else:
        critical_days = (
            np.argpartition(daily_totals, -2)[-2:]
            if daily_totals.size > 1
            else [int(np.argmax(daily_totals))]
        )
    
    peak_threshold = (
        np.percentile(hourly_totals[hourly_totals > 0], 75)
        if np.any(hourly_totals > 0)
        else 0
    )
    peak_hours = np.where(hourly_totals >= peak_threshold)[0]
    
    return {
        "daily_demand": daily_demand,
        "hourly_demand": hourly_demand,
        "active_days": active_days,
        "inactive_days": inactive_days,
        "working_days": working_days,
        "first_hour": first_hour,
        "last_hour": last_hour,
        "operating_hours": operating_hours,
        "peak_demand": peak_demand,
        "average_demand": avg_demand,
        "critical_days": critical_days,
        "peak_hours": peak_hours,
    }

def load_demand_matrix_from_df(df) -> np.ndarray:
    """Load demand matrix from DataFrame."""
    demand_matrix = np.zeros((7, 24), dtype=float)
    
    day_col = "Día"
    time_col = "Horario"
    demand_col = "Suma de Agentes Requeridos Erlang"
    
    if day_col not in df.columns or time_col not in df.columns or demand_col not in df.columns:
        for col in df.columns:
            if "día" in col.lower() or "dia" in col.lower():
                day_col = col
            elif "horario" in col.lower():
                time_col = col
            elif "erlang" in col.lower() or "requeridos" in col.lower():
                demand_col = col
    
    for _, row in df.iterrows():
        try:
            day = int(row[day_col])
            if not (1 <= day <= 7):
                continue
            day_idx = day - 1
            
            horario = str(row[time_col])
            if ":" in horario:
                hour = int(horario.split(":")[0])
            else:
                hour = int(float(horario))
            if not (0 <= hour <= 23):
                continue
            
            demanda = float(row[demand_col])
            demand_matrix[day_idx, hour] = demanda
        except (ValueError, TypeError, IndexError):
            continue
    
    return demand_matrix

def analyze_results(assignments, shifts_coverage, demand_matrix):
    """Analyze optimization results."""
    if not assignments:
        return None
    
    slots_per_day = len(next(iter(shifts_coverage.values()))) // 7 if shifts_coverage else 24
    total_coverage = np.zeros((7, slots_per_day), dtype=np.int32)
    total_agents = 0
    ft_agents = 0
    pt_agents = 0
    
    for shift_name, count in assignments.items():
        weekly_pattern = shifts_coverage[shift_name]
        slots_per_day = len(weekly_pattern) // 7
        pattern_matrix = np.array(weekly_pattern).reshape(7, slots_per_day)

        weekly_hours = pattern_matrix.sum()
        max_allowed = MAX_FT_HOURS if shift_name.startswith('FT') else MAX_PT_HOURS
        if weekly_hours > max_allowed:
            print(f"⚠️ {shift_name} exceeds maximum {max_allowed}h (has {weekly_hours}h)")
        
        total_coverage += pattern_matrix * count
        total_agents += count
        
        if shift_name.startswith('FT'):
            ft_agents += count
        else:
            pt_agents += count
    
    total_demand = demand_matrix.sum()
    total_covered = np.minimum(total_coverage, demand_matrix).sum()
    coverage_percentage = (total_covered / total_demand) * 100 if total_demand > 0 else 0
    
    diff_matrix = total_coverage - demand_matrix
    overstaffing = np.sum(diff_matrix[diff_matrix > 0])
    understaffing = np.sum(np.abs(diff_matrix[diff_matrix < 0]))
    
    return {
        'total_coverage': total_coverage,
        'total_agents': total_agents,
        'ft_agents': ft_agents,
        'pt_agents': pt_agents,
        'coverage_percentage': coverage_percentage,
        'overstaffing': overstaffing,
        'understaffing': understaffing,
        'diff_matrix': diff_matrix
    }

def create_heatmap(matrix, title, cmap='RdYlBu_r'):
    """Create heatmap visualization."""
    if plt is None:
        return None
    
    fig, ax = plt.subplots(figsize=(12, 6))
    im = ax.imshow(matrix, cmap=cmap, aspect='auto')
    
    ax.set_xticks(range(min(24, matrix.shape[1])))
    ax.set_xticklabels([f"{h:02d}" for h in range(min(24, matrix.shape[1]))])
    ax.set_yticks(range(7))
    ax.set_yticklabels(['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo'])
    
    for i in range(7):
        for j in range(min(24, matrix.shape[1])):
            if i < matrix.shape[0] and j < matrix.shape[1]:
                ax.text(j, i, f'{matrix[i, j]:.0f}',
                       ha="center", va="center", color="black", fontsize=8)
    
    ax.set_title(title)
    ax.set_xlabel('Hora del día')
    ax.set_ylabel('Día de la semana')
    plt.colorbar(im, ax=ax)
    return fig

def generate_all_heatmaps(demand, coverage=None, diff=None) -> dict:
    """Generate all heatmaps."""
    maps = {}
    if plt is not None:
        maps["demand"] = create_heatmap(demand, "Demanda por Hora y Día", "Reds")
        if coverage is not None:
            maps["coverage"] = create_heatmap(coverage, "Cobertura por Hora y Día", "Blues")
        if diff is not None:
            maps["difference"] = create_heatmap(diff, "Diferencias por Hora y Día", "RdBu")
    return maps

def single_model(func):
    """Ensure only one optimization model runs at a time."""
    @wraps(func)
    def wrapper(*args, **kwargs):
        with _MODEL_LOCK:
            return func(*args, **kwargs)
    return wrapper

@single_model
def optimize_with_greedy(shifts_coverage, demand_matrix, *, cfg=None):
    """Greedy optimization algorithm."""
    shifts_list = list(shifts_coverage.keys())
    assignments = {}
    current_coverage = np.zeros_like(demand_matrix, dtype=float)
    
    max_agents = max(100, int(demand_matrix.sum() / 12))
    
    for iteration in range(max_agents):
        best_shift = None
        best_score = -float("inf")
        best_pattern = None
        
        for shift_name in shifts_list:
            try:
                slots_per_day = len(shifts_coverage[shift_name]) // 7
                base_pattern = np.array(shifts_coverage[shift_name]).reshape(7, slots_per_day)
                
                if slots_per_day != demand_matrix.shape[1]:
                    pattern = np.zeros((7, demand_matrix.shape[1]))
                    cols_to_copy = min(slots_per_day, demand_matrix.shape[1])
                    pattern[:, :cols_to_copy] = base_pattern[:, :cols_to_copy]
                else:
                    pattern = base_pattern
                
                new_coverage = current_coverage + pattern
                
                current_deficit = np.maximum(0, demand_matrix - current_coverage)
                new_deficit = np.maximum(0, demand_matrix - new_coverage)
                deficit_reduction = np.sum(current_deficit - new_deficit)
                
                current_excess = np.maximum(0, current_coverage - demand_matrix)
                new_excess = np.maximum(0, new_coverage - demand_matrix)
                excess_increase = np.sum(new_excess - current_excess)
                
                score = deficit_reduction * 100 - excess_increase * 2.0
                
                if score > best_score:
                    best_score = score
                    best_shift = shift_name
                    best_pattern = pattern
                    
            except Exception as e:
                print(f"Error processing {shift_name}: {e}")
                continue
        
        if best_shift is None or best_score <= 0.5:
            break
        
        if best_shift not in assignments:
            assignments[best_shift] = 0
        assignments[best_shift] += 1
        current_coverage += best_pattern
        
        if np.sum(np.maximum(0, demand_matrix - current_coverage)) == 0:
            break
    
    return assignments, "GREEDY_ENHANCED"

def export_detailed_schedule(assignments, shifts_coverage):
    """Export detailed schedule to Excel."""
    if not assignments:
        return None
    
    try:
        from openpyxl import Workbook
    except ImportError:
        return None
    
    DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    
    wb = Workbook(write_only=True)
    ws_details = wb.create_sheet('Horarios_Semanales')
    ws_summary = wb.create_sheet('Resumen_Agentes')
    ws_shifts = wb.create_sheet('Turnos_Asignados')
    
    ws_details.append(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])
    ws_summary.append(['Agente', 'Turno', 'Dias_Trabajo'])
    ws_shifts.append(['Turno', 'Agentes'])
    
    summary_counts = defaultdict(int)
    agent_id = 1
    
    for shift_name, count in assignments.items():
        weekly_pattern = shifts_coverage[shift_name]
        slots_per_day = 24
        pattern_matrix = weekly_pattern.reshape(7, slots_per_day) if len(weekly_pattern) == 7 * slots_per_day else weekly_pattern.reshape(7, -1)[:, :slots_per_day]
        
        start_hour = 0.0
        for part in shift_name.split('_'):
            if '.' in part and part.replace('.', '').isdigit():
                try:
                    start_hour = float(part)
                    break
                except ValueError:
                    continue
        
        shift_type = 'FT' if shift_name.startswith('FT') else 'PT'
        
        for _ in range(count):
            for day in range(7):
                day_pattern = pattern_matrix[day]
                work_hours = np.where(day_pattern == 1)[0]
                
                if len(work_hours) > 0:
                    start_idx = int(start_hour)
                    end_idx = int(work_hours[-1]) + 1
                    next_day = end_idx <= start_idx
                    horario = f"{start_idx:02d}:00-{end_idx % 24:02d}:00" + ("+1" if next_day else "")
                    
                    if shift_name.startswith('FT'):
                        expected = list(range(start_idx, end_idx))
                        if next_day:
                            expected = list(range(start_idx, 24)) + list(range(0, end_idx % 24))
                        break_hours = set(expected) - set(work_hours)
                        if break_hours:
                            break_hour = min(break_hours)
                            break_time = f"{break_hour % 24:02d}:00-{((break_hour + 1) % 24) or 24:02d}:00"
                        else:
                            break_time = ""
                    else:
                        break_time = ""
                    
                    row = [f"AGT_{agent_id:03d}", DAYS[day], horario, break_time, shift_name, shift_type]
                else:
                    row = [f"AGT_{agent_id:03d}", DAYS[day], "DSO", "", shift_name, "DSO"]
                
                ws_details.append(row)
                summary_counts[(agent_id, shift_name)] += 1
            agent_id += 1
    
    for (agent_idx, shift_name), cnt in summary_counts.items():
        ws_summary.append([f"AGT_{agent_idx:03d}", shift_name, cnt])
    
    for shift, count in assignments.items():
        ws_shifts.append([shift, count])
    
    excel_io = BytesIO()
    wb.save(excel_io)
    return excel_io.getvalue()

# Default configuration
DEFAULT_CONFIG = {
    "solver_time": 300,
    "solver_msg": 1,
    "TARGET_COVERAGE": 98.0,
    "agent_limit_factor": 12,
    "excess_penalty": 2.0,
    "peak_bonus": 1.5,
    "critical_bonus": 2.0,
    "iterations": 30,
    "use_ft": True,
    "use_pt": True,
    "allow_8h": True,
    "allow_10h8": False,
    "allow_pt_4h": True,
    "allow_pt_6h": True,
    "allow_pt_5h": False,
    "break_from_start": 2.5,
    "break_from_end": 2.5,
    "ACTIVE_DAYS": list(range(7)),
}

def merge_config(cfg=None):
    """Merge configuration with defaults."""
    merged = DEFAULT_CONFIG.copy()
    if cfg:
        merged.update({k: v for k, v in cfg.items() if v is not None})
    return merged
