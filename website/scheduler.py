# -*- coding: utf-8 -*-
from flask import current_app
import json
import time
import os
import gc
import hashlib
import signal
from io import BytesIO, StringIO
from itertools import combinations, permutations
import heapq

import tempfile
import csv

import numpy as np
try:
    import matplotlib
    matplotlib.use("Agg")  # Use a non-GUI backend for server-side generation
    import matplotlib.pyplot as plt
    import seaborn as sns
except Exception:  # pragma: no cover - optional dependency
    matplotlib = None
    plt = None
    sns = None
try:
    import psutil
except Exception:  # pragma: no cover - optional dependency
    psutil = None

# Lookup table for population count and global context
POP = np.fromiter((bin(i).count("1") for i in range(256)), dtype=np.uint8)
CONTEXT = {}


def unpack_pattern(packed: np.ndarray, cols: int) -> np.ndarray:
    """Unpack a packed weekly pattern to shape (7, cols)."""
    return np.unpackbits(packed.reshape(7, -1), axis=1)[:, :cols]

try:
    import pulp as pl
    PULP_AVAILABLE = True
except Exception:
    PULP_AVAILABLE = False

print(f"[OPTIMIZER] PuLP disponible: {PULP_AVAILABLE}")

from threading import RLock, current_thread
from functools import wraps
import ctypes

_MODEL_LOCK = RLock()

# Registry of active optimization jobs
active_jobs = {}


def init_app(app):
    """Initialize scheduler storage and defaults on ``app``."""

    if not hasattr(app, "extensions"):
        app.extensions = {}
    if "scheduler" not in app.extensions:
        app.extensions["scheduler"] = {"jobs": {}, "results": {}, "active_jobs": active_jobs}
    app.config.setdefault("SCHEDULER_MAX_RUNTIME", 300)
    app.config.setdefault("SCHEDULER_MAX_MEMORY_GB", None)


def _store(app=None):
    """Return the shared scheduler store."""

    app = app or current_app
    if not hasattr(app, "extensions"):
        app.extensions = {}
    return app.extensions.setdefault("scheduler", {"jobs": {}, "results": {}, "active_jobs": active_jobs})


def mark_running(job_id, app=None):
    s = _store(app)
    s.setdefault("jobs", {})[job_id] = {"status": "running", "progress": {}}


def mark_finished(job_id, result, excel_path, csv_path, app=None):
    s = _store(app)
    s.setdefault("jobs", {})[job_id] = {"status": "finished"}
    s.setdefault("results", {})[job_id] = {
        "result": result,
        "excel_path": excel_path,
        "csv_path": csv_path,
        "timestamp": time.time(),
    }


def mark_error(job_id, error_msg, app=None):
    s = _store(app)
    s.setdefault("jobs", {})[job_id] = {"status": "error", "error": error_msg}


def mark_cancelled(job_id, app=None):
    s = _store(app)
    s.setdefault("jobs", {})[job_id] = {"status": "cancelled"}


def get_status(job_id, app=None):
    return _store(app).get("jobs", {}).get(job_id, {"status": "unknown", "progress": {}})

def update_progress(job_id, progress_info, app=None):
    """Actualizar información de progreso de un job."""
    s = _store(app)
    job_data = s.setdefault("jobs", {}).setdefault(job_id, {"status": "running", "progress": {}})
    job_data["progress"].update(progress_info)


def get_result(job_id, app=None):
    return _store(app).get("results", {}).get(job_id)


# Compatibility alias
get_payload = get_result


def _stop_thread(thread):
    """Attempt to stop a running thread by raising ``SystemExit`` inside it."""
    if thread and thread.is_alive():  # pragma: no cover - safety guard
        ctypes.pythonapi.PyThreadState_SetAsyncExc(
            ctypes.c_long(thread.ident), ctypes.py_object(SystemExit)
        )


def single_model(func):
    """Ensure that only one optimization model is built/solved at a time."""
    @wraps(func)
    def wrapper(*args, **kwargs):
        with _MODEL_LOCK:
            return func(*args, **kwargs)

    return wrapper

# Default configuration values used when no override is supplied
DEFAULT_CONFIG = {
    # Streamlit legacy defaults from ``legacy/app1.py``
    "solver_time": 240,  # EXACTO del legacy
    "solver_msg": 1,
    "TARGET_COVERAGE": 98.0,
    "agent_limit_factor": 12,
    "excess_penalty": 2.0,
    "peak_bonus": 1.5,
    "critical_bonus": 2.0,
    "iterations": 30,
    "solver_threads": os.cpu_count() or 1,
    "max_memory_gb": None,  # None uses all available memory
    "max_patterns": None,
    "batch_size": 2000,
    "quality_threshold": 0,
    "use_ft": True,
    "use_pt": True,
    "allow_8h": True,
    "allow_10h8": False,
    "allow_pt_4h": True,
    "allow_pt_6h": True,
    "allow_pt_5h": False,
    "break_from_start": 2.5,
    "break_from_end": 2.5,
    "optimization_profile": "Equilibrado (Recomendado)",
    "profile": "Equilibrado (Recomendado)",  # Alias para compatibilidad
    "ACTIVE_DAYS": list(range(7)),
    "K": 1000,
}


def merge_config(cfg=None):
    """Return a configuration dictionary overlaying defaults with ``cfg``."""
    merged = DEFAULT_CONFIG.copy()
    try:
        app_cfg = current_app.config
        merged.update({k: app_cfg[k] for k in DEFAULT_CONFIG.keys() if k in app_cfg})
    except Exception:
        pass
    if cfg:
        merged.update({k: v for k, v in cfg.items() if v is not None})
    return merged


def _build_pattern(days, durations, start_hour, break_len, break_from_start,
                   break_from_end, slot_factor=1):
    """Construct a weekly pattern array.

    ``days`` and ``durations`` must align. A break of ``break_len`` hours is
    placed between ``break_from_start`` and ``break_from_end`` from the shift
    start. ``slot_factor`` controls the number of slots per hour.

    Returns a flattened :class:`numpy.ndarray` of shape ``7 * slots_per_day``.
    """
    slots_per_day = 24 * slot_factor
    pattern = np.zeros((7, slots_per_day), dtype=np.int8)
    for day, dur in zip(days, durations):
        for s in range(int(dur * slot_factor)):
            slot = int(start_hour * slot_factor) + s
            d_off, idx = divmod(slot, slots_per_day)
            pattern[(day + d_off) % 7, idx] = 1
        if break_len:
            b_start = int((start_hour + break_from_start) * slot_factor)
            b_end = int((start_hour + dur - break_from_end) * slot_factor)
            if b_start < b_end:
                b_slot = b_start + (b_end - b_start) // 2
            else:
                b_slot = b_start
            for b in range(int(break_len * slot_factor)):
                slot = b_slot + b
                d_off, idx = divmod(slot, slots_per_day)
                pattern[(day + d_off) % 7, idx] = 0
    return pattern.flatten()


def memory_limit_patterns(slots_per_day, max_gb=None):
    """Return the max number of patterns that fit in memory.

    ``max_gb`` caps the available memory (in gigabytes) used for the
    calculation when provided. When ``max_gb`` is ``None`` all available
    memory is considered.
    """
    if slots_per_day <= 0:
        return 0
    if psutil is None:
        if max_gb is None:
            raise RuntimeError("psutil is required to determine available memory")
        cap = max_gb * 1024 ** 3
    else:
        available = psutil.virtual_memory().available
        if max_gb is not None:
            cap = min(available, max_gb * 1024 ** 3)
        else:
            cap = available
    return int(cap // (7 * slots_per_day))


def monitor_memory_usage():
    """Return current memory usage percentage."""
    if psutil is None:
        return 0.0
    return psutil.virtual_memory().percent


def adaptive_chunk_size(base=5000):
    """Adjust chunk size based on memory usage."""
    if psutil is None:
        return base
    usage = monitor_memory_usage()
    if usage > 80:
        return max(1000, base // 4)
    if usage > 60:
        return max(2000, base // 2)
    return base


def emergency_cleanup(threshold=85.0):
    """Trigger ``gc.collect`` if usage exceeds ``threshold``."""
    if psutil is None:
        return False
    if monitor_memory_usage() >= threshold:
        gc.collect()
        return True
    return False


def get_smart_start_hours(demand_matrix, max_hours=12):
    """Return a list of start hours around peak demand."""
    if demand_matrix is None or demand_matrix.size == 0:
        return [float(h) for h in range(24)]

    cols = demand_matrix.shape[1]
    hourly_totals = demand_matrix.sum(axis=0)
    top = np.argsort(hourly_totals)[-max_hours:]
    hours = sorted({round(h / cols * 24, 2) for h in top})
    return hours


def score_pattern(pattern: np.ndarray, demand_matrix: np.ndarray) -> int:
    """Quick heuristic score to sort patterns before solving."""
    dm = demand_matrix.flatten()
    pat = pattern.astype(int)
    lim = min(len(dm), len(pat))
    return int(np.minimum(pat[:lim], dm[:lim]).sum())


def _resize_matrix(matrix, target_cols):
    """Resize ``matrix`` to ``target_cols`` via repeat or max pooling."""
    if matrix.shape[1] == target_cols:
        return matrix
    if matrix.shape[1] < target_cols:
        factor = target_cols // matrix.shape[1]
        return np.repeat(matrix, factor, axis=1)
    factor = matrix.shape[1] // target_cols
    return matrix.reshape(matrix.shape[0], target_cols, factor).max(axis=2)


def score_and_filter_patterns(patterns, demand_matrix, *, keep_percentage=0.3,
                              peak_bonus=1.5, critical_bonus=2.0,
                              efficiency_bonus=1.0):
    """Score patterns against demand and keep the best ones."""
    if demand_matrix is None or not patterns:
        return patterns
    dm = np.asarray(demand_matrix, dtype=float)
    days, hours = dm.shape
    daily_totals = dm.sum(axis=1)
    hourly_totals = dm.sum(axis=0)
    critical_days = (np.argsort(daily_totals)[-2:]
                     if daily_totals.size > 1 else [int(np.argmax(daily_totals))])
    if np.any(hourly_totals > 0):
        thresh = np.percentile(hourly_totals[hourly_totals > 0], 75)
        peak_hours = np.where(hourly_totals >= thresh)[0]
    else:
        peak_hours = []
    scores = []
    for name, pat in patterns.items():
        cols = len(pat) // 7
        pat_mat = pat.reshape(7, cols)
        dm_resized = _resize_matrix(dm, cols)
        coverage = np.minimum(pat_mat, dm_resized)
        score = coverage.sum()
        total_hours = pat_mat.sum()
        if total_hours > 0:
            efficiency = coverage.sum() / total_hours
            score += efficiency * efficiency_bonus
        if len(critical_days) > 0:
            score += coverage[critical_days].sum() * critical_bonus
        if len(peak_hours) > 0:
            ph = [h for h in peak_hours if h < cols]
            if ph:
                score += coverage[:, ph].sum() * peak_bonus
        scores.append((name, score))
    scores.sort(key=lambda x: x[1], reverse=True)
    keep_n = max(1, int(len(scores) * keep_percentage))
    top = {name for name, _ in scores[:keep_n]}
    return {k: patterns[k] for k in top}


def load_shift_patterns(cfg, *, start_hours=None, break_from_start=2.0,
                         break_from_end=2.0, slot_duration_minutes=30,
                         max_patterns=None, demand_matrix=None,
                         keep_percentage=0.3, peak_bonus=1.5,
                         critical_bonus=2.0, efficiency_bonus=1.0,
                         max_patterns_per_shift=None, smart_start_hours=False):
    """Load shift patterns from ``cfg`` and return them as a dict."""
    if isinstance(cfg, str):
        with open(cfg, "r") as fh:
            data = json.load(fh)
    else:
        data = cfg
    if slot_duration_minutes is not None and 60 % slot_duration_minutes != 0:
        raise ValueError("slot_duration_minutes must divide 60")
    base_slot_min = slot_duration_minutes or 60
    slots_per_day = 24 * (60 // base_slot_min)
    if max_patterns is None:
        max_gb = cfg.get("max_memory_gb")
        if isinstance(max_gb, str) and not max_gb.strip():
            max_gb = None
        max_patterns = memory_limit_patterns(slots_per_day, max_gb=max_gb)
    shifts_coverage = {}
    unique_patterns = {}
    for shift in data.get("shifts", []):
        name = shift.get("name", "SHIFT")
        pat = shift.get("pattern", {})
        brk = shift.get("break", 0)
        slot_min = slot_duration_minutes or shift.get("slot_duration_minutes", 60)
        if 60 % slot_min != 0:
            raise ValueError("slot_duration_minutes must divide 60")
        step = slot_min / 60
        slot_factor = 60 // slot_min
        base_hours = list(start_hours) if start_hours is not None else list(np.arange(0, 24, step))
        if smart_start_hours and demand_matrix is not None:
            smart = get_smart_start_hours(demand_matrix)
            sh_hours = [h for h in base_hours if any(abs(h - s) < step/2 for s in smart)]
        else:
            sh_hours = base_hours
        work_days = pat.get("work_days", [])
        segments_spec = pat.get("segments", [])
        segments = []
        for seg in segments_spec:
            if isinstance(seg, dict):
                hours = seg.get("hours")
                count = seg.get("count", 1)
                if hours is None:
                    continue
                segments.extend([int(hours)] * int(count))
            else:
                segments.append(int(seg))
        if isinstance(work_days, int):
            day_candidates = range(7)
            day_combos = combinations(day_candidates, work_days)
        else:
            day_combos = combinations(work_days, min(len(segments), len(work_days)))
        if isinstance(brk, dict):
            if brk.get("enabled", False):
                brk_len = brk.get("length_minutes", 0) / 60
                brk_start = brk.get("earliest_after_start", 0) / 60
                brk_end = brk.get("latest_before_end", 0) / 60
            else:
                brk_len = 0
                brk_start = break_from_start
                brk_end = break_from_end
        else:
            brk_len = float(brk)
            brk_start = break_from_start
            brk_end = break_from_end
        shift_patterns = {}
        for days_sel in day_combos:
            for perm in set(permutations(segments, len(days_sel))):
                for sh in sh_hours:
                    pattern = _build_pattern(days_sel, perm, sh, brk_len, brk_start, brk_end, slot_factor)
                    pat_key = pattern.tobytes()
                    if pat_key in unique_patterns:
                        continue
                    day_str = "".join(map(str, days_sel))
                    seg_str = "_".join(map(str, perm))
                    shift_name = f"{name}_{sh:04.1f}_{day_str}_{seg_str}"
                    shift_patterns[shift_name] = pattern
                    unique_patterns[pat_key] = shift_name
                    if max_patterns_per_shift and len(shift_patterns) >= max_patterns_per_shift:
                        break
                    if max_patterns is not None and len(shifts_coverage) + len(shift_patterns) >= max_patterns:
                        break
                if max_patterns_per_shift and len(shift_patterns) >= max_patterns_per_shift:
                    break
                if max_patterns is not None and len(shifts_coverage) + len(shift_patterns) >= max_patterns:
                    break
            if max_patterns_per_shift and len(shift_patterns) >= max_patterns_per_shift:
                break
        if demand_matrix is not None:
            shift_patterns = score_and_filter_patterns(
                shift_patterns,
                demand_matrix,
                keep_percentage=keep_percentage,
                peak_bonus=peak_bonus,
                critical_bonus=critical_bonus,
                efficiency_bonus=efficiency_bonus,
            )
        shifts_coverage.update(shift_patterns)
        monitor_memory_usage()
        gc.collect()
        if max_patterns is not None and len(shifts_coverage) >= max_patterns:
            return shifts_coverage
    if demand_matrix is not None:
        shifts_coverage = score_and_filter_patterns(
            shifts_coverage,
            demand_matrix,
            keep_percentage=keep_percentage,
            peak_bonus=peak_bonus,
            critical_bonus=critical_bonus,
            efficiency_bonus=efficiency_bonus,
        )
    return shifts_coverage


def load_learning_data():
    """Load optimization learning data from disk if available."""
    try:
        if os.path.exists("optimization_learning.json"):
            with open("optimization_learning.json", "r") as f:
                return json.load(f)
    except Exception:
        pass
    return {"executions": [], "best_params": {}, "stats": {}}


def save_learning_data(data):
    """Persist optimization learning data to disk."""
    try:
        with open("optimization_learning.json", "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


def get_adaptive_params(demand_matrix, target_coverage):
    """Return tuned parameters based on previous executions."""
    learning_data = load_learning_data()
    input_hash = hashlib.md5(str(demand_matrix).encode()).hexdigest()[:12]
    total_demand = demand_matrix.sum()
    peak_demand = demand_matrix.max()
    similar_runs = [e for e in learning_data.get("executions", []) if e.get("input_hash") == input_hash]
    similar_runs.sort(key=lambda x: x.get("timestamp", 0))
    if similar_runs:
        last_run = similar_runs[-1]
        best_run = max(similar_runs, key=lambda x: x.get("coverage", 0))
        last_cov = last_run.get("coverage", 0)
        coverage_gap = target_coverage - last_cov
        base_params = best_run.get("params", {})
        evolution_factor = min(0.3, len(similar_runs) * 0.05)
        if len(similar_runs) >= 3:
            recent_coverages = [run.get("coverage", 0) for run in similar_runs[-3:]]
            if recent_coverages[-1] <= recent_coverages[-2]:
                evolution_factor *= 2
        if coverage_gap > 10:
            return {
                "agent_limit_factor": max(5, int(base_params.get("agent_limit_factor", 20) * (1 - evolution_factor))),
                "excess_penalty": base_params.get("excess_penalty", 0.1) * (1 - evolution_factor),
                "peak_bonus": base_params.get("peak_bonus", 2.0) * (1 + evolution_factor),
                "critical_bonus": base_params.get("critical_bonus", 2.5) * (1 + evolution_factor),
                "precision_mode": True,
                "learned": True,
                "runs_count": len(similar_runs),
                "evolution_step": "aggressive",
            }
        elif coverage_gap > 3:
            return {
                "agent_limit_factor": max(8, int(base_params.get("agent_limit_factor", 20) * (1 - evolution_factor * 0.5))),
                "excess_penalty": base_params.get("excess_penalty", 0.2) * (1 - evolution_factor * 0.5),
                "peak_bonus": base_params.get("peak_bonus", 1.8) * (1 + evolution_factor * 0.5),
                "critical_bonus": base_params.get("critical_bonus", 2.0) * (1 + evolution_factor * 0.5),
                "precision_mode": True,
                "learned": True,
                "runs_count": len(similar_runs),
                "evolution_step": "moderate",
            }
        elif coverage_gap > 0:
            return {
                "agent_limit_factor": max(12, int(base_params.get("agent_limit_factor", 22) * (1 - evolution_factor * 0.2))),
                "excess_penalty": base_params.get("excess_penalty", 0.3) * (1 - evolution_factor * 0.2),
                "peak_bonus": base_params.get("peak_bonus", 1.5) * (1 + evolution_factor * 0.2),
                "critical_bonus": base_params.get("critical_bonus", 1.8) * (1 + evolution_factor * 0.2),
                "precision_mode": False,
                "learned": True,
                "runs_count": len(similar_runs),
                "evolution_step": "fine_tune",
            }
        else:
            noise = np.random.uniform(-0.1, 0.1)
            return {
                "agent_limit_factor": max(15, int(base_params.get("agent_limit_factor", 20) * (1 + noise))),
                "excess_penalty": max(0.01, base_params.get("excess_penalty", 0.5) * (1 + noise)),
                "peak_bonus": base_params.get("peak_bonus", 1.5) * (1 + noise * 0.5),
                "critical_bonus": base_params.get("critical_bonus", 2.0) * (1 + noise * 0.5),
                "precision_mode": False,
                "learned": True,
                "runs_count": len(similar_runs),
                "evolution_step": "explore",
            }
    return {
        "agent_limit_factor": max(8, int(total_demand / max(1, peak_demand) * 3)),
        "excess_penalty": 0.05,
        "peak_bonus": 2.5,
        "critical_bonus": 3.0,
        "precision_mode": True,
        "learned": False,
        "runs_count": 0,
        "evolution_step": "initial",
    }


def save_execution_result(demand_matrix, params, coverage, total_agents, execution_time):
    """Record an execution summary in the learning file."""
    learning_data = load_learning_data()
    input_hash = hashlib.md5(str(demand_matrix).encode()).hexdigest()[:12]
    efficiency_score = coverage / max(1, total_agents * 0.1)
    balance_score = coverage - abs(coverage - 100) * 0.5
    execution_result = {
        "timestamp": time.time(),
        "input_hash": input_hash,
        "params": {
            "agent_limit_factor": params.get("agent_limit_factor"),
            "excess_penalty": params.get("excess_penalty"),
            "peak_bonus": params.get("peak_bonus"),
            "critical_bonus": params.get("critical_bonus"),
        },
        "coverage": coverage,
        "total_agents": total_agents,
        "efficiency_score": efficiency_score,
        "balance_score": balance_score,
        "execution_time": execution_time,
        "demand_total": float(demand_matrix.sum()),
        "evolution_step": params.get("evolution_step", "unknown"),
    }
    learning_data.setdefault("executions", []).append(execution_result)
    pattern_executions = [e for e in learning_data["executions"] if e.get("input_hash") == input_hash]
    if len(pattern_executions) > 50:
        learning_data["executions"] = [e for e in learning_data["executions"] if e.get("input_hash") != input_hash or e.get("timestamp", 0) >= sorted([p.get("timestamp", 0) for p in pattern_executions])[-50]]
    current_best = learning_data.get("best_params", {}).get(input_hash, {})
    new_score = efficiency_score if coverage >= 98 else coverage * 2
    if not current_best or new_score > current_best.get("score", 0):
        learning_data.setdefault("best_params", {})[input_hash] = {
            "params": execution_result["params"],
            "coverage": coverage,
            "total_agents": total_agents,
            "score": new_score,
            "efficiency_score": efficiency_score,
            "timestamp": time.time(),
        }
    pattern_runs = [e for e in learning_data["executions"] if e.get("input_hash") == input_hash]
    if len(pattern_runs) >= 2:
        recent_improvement = pattern_runs[-1]["coverage"] - pattern_runs[-2]["coverage"]
    else:
        recent_improvement = 0
    learning_data["stats"] = {
        "total_executions": len(learning_data["executions"]),
        "unique_patterns": len(set(e["input_hash"] for e in learning_data["executions"])),
        "avg_coverage": np.mean([e["coverage"] for e in learning_data["executions"][-10:]]),
        "recent_improvement": recent_improvement,
        "best_coverage": max([e["coverage"] for e in learning_data["executions"]], default=0),
        "last_updated": time.time(),
    }
    save_learning_data(learning_data)
    return True

# ---------------------------------------------------------------------------
# Demand utilities
# ---------------------------------------------------------------------------

def load_demand_excel(file_stream) -> np.ndarray:
    """Parse an Excel file exported from Ntech and return a matrix."""
    import pandas as pd

    df = pd.read_excel(file_stream)
    day_col = [c for c in df.columns if "Día" in c][0]
    demand_col = [c for c in df.columns if "Erlang" in c or "Requeridos" in c][-1]
    dm = df.pivot_table(index=day_col, values=demand_col, columns=df.index % 24, aggfunc="first").fillna(0)
    dm = dm.reindex(range(1, 8)).fillna(0)
    dm = dm.sort_index()
    matrix = dm.to_numpy(dtype=int)
    if matrix.shape[1] != 24:
        matrix = np.pad(matrix, ((0, 0), (0, 24 - matrix.shape[1])), constant_values=0)
    return matrix


def heatmap(matrix: np.ndarray, title: str) -> BytesIO:
    """Return an image buffer with ``matrix`` rendered as a heatmap."""
    fig, ax = plt.subplots(figsize=(10, 4))
    sns.heatmap(matrix, ax=ax, cmap="viridis", cbar=False)
    ax.set_title(title)
    ax.set_xlabel("Hora")
    ax.set_ylabel("Dia")
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def load_demand_matrix_from_df(df) -> np.ndarray:
    """Return a 7x24 demand matrix from an Ntech formatted dataframe."""

    demand_matrix = np.zeros((7, 24), dtype=float)

    day_col = "Día"
    time_col = "Horario"
    demand_col = "Suma de Agentes Requeridos Erlang"

    if day_col not in df.columns or time_col not in df.columns or demand_col not in df.columns:
        for col in df.columns:
            if "día" in col.lower() or "dia" in col.lower():
                day_col = col
            elif "horario" in col.lower():
                time_col = col
            elif "erlang" in col.lower() or "requeridos" in col.lower():
                demand_col = col

    for _, row in df.iterrows():
        try:
            day = int(row[day_col])
            if not (1 <= day <= 7):
                continue
            day_idx = day - 1

            horario = str(row[time_col])
            if ":" in horario:
                hour = int(horario.split(":")[0])
            else:
                hour = int(float(horario))
            if not (0 <= hour <= 23):
                continue

            demanda = float(row[demand_col])
            demand_matrix[day_idx, hour] = demanda
        except (ValueError, TypeError, IndexError):
            continue

    return demand_matrix


def analyze_demand_matrix(matrix: np.ndarray) -> dict:
    """Return basic metrics from a demand matrix."""
    daily_demand = matrix.sum(axis=1)
    hourly_demand = matrix.sum(axis=0)
    active_days = [d for d in range(7) if daily_demand[d] > 0]
    inactive_days = [d for d in range(7) if daily_demand[d] == 0]
    working_days = len(active_days)
    active_hours = np.where(hourly_demand > 0)[0]
    first_hour = int(active_hours.min()) if active_hours.size else 8
    last_hour = int(active_hours.max()) if active_hours.size else 20
    operating_hours = last_hour - first_hour + 1
    peak_demand = float(matrix.max()) if matrix.size else 0.0
    avg_demand = float(matrix[active_days].mean()) if active_days else 0.0
    daily_totals = matrix.sum(axis=1)
    hourly_totals = matrix.sum(axis=0)
    critical_days = (
        np.argsort(daily_totals)[-2:] if daily_totals.size > 1 else [int(np.argmax(daily_totals))]
    )
    peak_threshold = (
        np.percentile(hourly_totals[hourly_totals > 0], 75)
        if np.any(hourly_totals > 0)
        else 0
    )
    peak_hours = np.where(hourly_totals >= peak_threshold)[0]
    return {
        "daily_demand": daily_demand,
        "hourly_demand": hourly_demand,
        "active_days": active_days,
        "inactive_days": inactive_days,
        "working_days": working_days,
        "first_hour": first_hour,
        "last_hour": last_hour,
        "operating_hours": operating_hours,
        "peak_demand": peak_demand,
        "average_demand": avg_demand,
        "critical_days": critical_days,
        "peak_hours": peak_hours,
    }


def create_heatmap(matrix, title, cmap="RdYlBu_r"):
    """Return a matplotlib figure with ``matrix`` visualised as a heatmap."""
    fig, ax = plt.subplots(figsize=(12, 6))
    im = ax.imshow(matrix, cmap=cmap, aspect="auto")
    ax.set_xticks(range(24))
    ax.set_xticklabels([f"{h:02d}" for h in range(24)])
    ax.set_yticks(range(7))
    ax.set_yticklabels([
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    ])
    for i in range(7):
        for j in range(24):
            ax.text(j, i, f"{matrix[i, j]:.0f}", ha="center", va="center", color="black", fontsize=8)
    ax.set_title(title)
    ax.set_xlabel("Hora del día")
    ax.set_ylabel("Día de la semana")
    plt.colorbar(im, ax=ax)
    return fig


def generate_all_heatmaps(demand, coverage=None, diff=None) -> dict:
    """Generate heatmaps for demand, coverage and difference matrices."""
    maps = {"demand": create_heatmap(demand, "Demanda por Hora y Día", "Reds")}
    if coverage is not None:
        maps["coverage"] = create_heatmap(coverage, "Cobertura por Hora y Día", "Blues")
    if diff is not None:
        maps["difference"] = create_heatmap(diff, "Diferencias por Hora y Día", "RdBu")
    return maps


# Perfiles EXACTOS del legacy Streamlit
PROFILES = {
    "Equilibrado (Recomendado)": {
        "agent_limit_factor": 12,
        "excess_penalty": 2.0,
        "peak_bonus": 1.5,
        "critical_bonus": 2.0,
        "description": "Balance óptimo entre cobertura y costo",
        "strategy": "balanced",
        "solver_time": 300,
        "precision_mode": False,
    },
    "Conservador": {
        "agent_limit_factor": 30,
        "excess_penalty": 0.5,
        "peak_bonus": 1.0,
        "critical_bonus": 1.2,
        "description": "Minimiza riesgos, permite más agentes",
        "strategy": "conservative",
        "solver_time": 240,
        "precision_mode": False,
    },
    "Agresivo": {
        "agent_limit_factor": 15,
        "excess_penalty": 0.05,
        "peak_bonus": 1.5,
        "critical_bonus": 2.0,
        "description": "Maximiza eficiencia, tolera déficit menor",
        "strategy": "aggressive",
        "solver_time": 180,
        "precision_mode": True,
    },
    "Máxima Cobertura": {
        "agent_limit_factor": 7,
        "excess_penalty": 0.005,
        "peak_bonus": 3.0,
        "critical_bonus": 4.0,
        "description": "Prioriza cobertura completa sobre costo",
        "strategy": "max_coverage",
        "solver_time": 400,
        "precision_mode": True,
        "target_coverage": 99.5,
    },
    "Mínimo Costo": {
        "agent_limit_factor": 35,
        "excess_penalty": 0.8,
        "peak_bonus": 0.8,
        "critical_bonus": 1.0,
        "description": "Minimiza número de agentes",
        "strategy": "min_cost",
        "solver_time": 200,
        "precision_mode": False,
        "allow_deficit": True,
    },
    "100% Cobertura Eficiente": {
        "agent_limit_factor": 6,
        "excess_penalty": 0.01,
        "peak_bonus": 3.5,
        "critical_bonus": 4.5,
        "description": "Cobertura completa con mínimo exceso",
        "strategy": "perfect_efficient",
        "solver_time": 500,
        "precision_mode": True,
        "target_coverage": 100.0,
        "max_excess_ratio": 0.02,
    },
    "100% Cobertura Total": {
        "agent_limit_factor": 5,
        "excess_penalty": 0.001,
        "peak_bonus": 4.0,
        "critical_bonus": 5.0,
        "description": "Cobertura completa sin restricciones de exceso",
        "strategy": "perfect_total",
        "solver_time": 600,
        "precision_mode": True,
        "target_coverage": 100.0,
        "allow_excess": True,
    },
    "Cobertura Perfecta": {
        "agent_limit_factor": 8,
        "excess_penalty": 0.01,
        "peak_bonus": 3.0,
        "critical_bonus": 4.0,
        "description": "Balance entre cobertura perfecta y eficiencia",
        "strategy": "perfect_balanced",
        "solver_time": 450,
        "precision_mode": True,
        "target_coverage": 99.8,
    },
    "100% Exacto": {
        "agent_limit_factor": 6,
        "excess_penalty": 0.005,
        "peak_bonus": 4.0,
        "critical_bonus": 5.0,
        "description": "Cobertura exacta sin déficit ni exceso",
        "strategy": "exact_match",
        "solver_time": 600,
        "precision_mode": True,
        "target_coverage": 100.0,
        "zero_excess": True,
        "zero_deficit": True,
    },
    "JEAN": {
        "agent_limit_factor": 30,
        "excess_penalty": 5.0,
        "peak_bonus": 2.0,
        "critical_bonus": 2.5,
        "TARGET_COVERAGE": 98.0,
        "description": "Búsqueda iterativa para minimizar exceso+déficit",
        "strategy": "jean_search",
        "solver_time": 240,  # EXACTO del legacy
        "precision_mode": True,
        "iterative_search": True,
        "search_iterations": 5,
        "use_jean_search": True,
    },
    "JEAN Personalizado": {
        "agent_limit_factor": 30,
        "excess_penalty": 5.0,
        "peak_bonus": 2.0,
        "critical_bonus": 2.5,
        "TARGET_COVERAGE": 98.0,
        "description": "JEAN con configuración personalizada de turnos",
        "strategy": "jean_custom",
        "solver_time": 240,  # EXACTO del legacy
        "precision_mode": True,
        "custom_shifts": True,
        "ft_pt_strategy": True,
        "use_jean_search": True,
        "slot_duration_minutes": 30,
    },
    "Personalizado": {
        "agent_limit_factor": 25,
        "excess_penalty": 0.5,
        "peak_bonus": 1.5,
        "critical_bonus": 2.0,
        "description": "Configuración manual de parámetros",
        "strategy": "custom",
        "solver_time": 300,
        "precision_mode": False,
        "user_defined": True,
    },
    "Aprendizaje Adaptativo": {
        "agent_limit_factor": 8,
        "excess_penalty": 0.01,
        "peak_bonus": 3.0,
        "critical_bonus": 4.0,
        "description": "IA que evoluciona automáticamente",
        "strategy": "adaptive_learning",
        "solver_time": 350,
        "precision_mode": True,
        "adaptive": True,
        "learning_enabled": True,
    },
}


def apply_configuration(cfg=None):
    """Apply an optimization profile over ``cfg`` and return the result."""
    cfg = merge_config(cfg)
    profile = cfg.get("optimization_profile", "Equilibrado (Recomendado)")
    
    print(f"[CONFIG] Aplicando perfil: {profile}")
    
    # Obtener parámetros del perfil
    profile_params = PROFILES.get(profile)
    
    if profile_params:
        # Aplicar todos los parámetros del perfil
        for key, val in profile_params.items():
            if key not in cfg or cfg[key] is None:
                cfg[key] = val
        
        # Configuraciones específicas por perfil
        if profile == "JEAN":
            cfg["optimization_profile"] = "JEAN"
            cfg["use_jean_search"] = True
            print(f"[CONFIG] JEAN: factor={cfg['agent_limit_factor']}, penalty={cfg['excess_penalty']}, target={cfg.get('TARGET_COVERAGE', 98)}%")
        
        elif profile == "JEAN Personalizado":
            cfg["optimization_profile"] = "JEAN Personalizado"
            cfg["use_jean_search"] = True
            # Solo habilitar custom shifts si hay JSON
            if cfg.get("custom_shifts_json"):
                cfg["use_custom_shifts"] = True
                print(f"[CONFIG] JEAN Personalizado: shifts personalizados habilitados")
            else:
                print(f"[CONFIG] JEAN Personalizado: usando lógica JEAN estándar (sin JSON personalizado)")
        
        elif profile == "Aprendizaje Adaptativo":
            # Aplicar parámetros adaptativos si están disponibles
            if cfg.get("demand_matrix") is not None:
                adaptive_params = get_adaptive_params(cfg["demand_matrix"], cfg.get("TARGET_COVERAGE", 98.0))
                cfg.update(adaptive_params)
                print(f"[CONFIG] Adaptativo: parámetros evolutivos aplicados")
        
        elif profile in ["100% Cobertura Eficiente", "100% Cobertura Total", "100% Exacto"]:
            cfg["use_perfect_coverage"] = True
            cfg["strict_coverage"] = True
            print(f"[CONFIG] {profile}: modo cobertura perfecta")
        
        elif profile == "Máxima Cobertura":
            cfg["prioritize_coverage"] = True
            cfg["allow_high_agents"] = True
            print(f"[CONFIG] Máxima Cobertura: prioridad en cobertura completa")
        
        elif profile == "Mínimo Costo":
            cfg["minimize_agents"] = True
            cfg["cost_priority"] = True
            print(f"[CONFIG] Mínimo Costo: minimización de agentes")
        
        elif profile == "Agresivo":
            cfg["aggressive_optimization"] = True
            cfg["fast_solve"] = True
            print(f"[CONFIG] Agresivo: optimización agresiva")
        
        elif profile == "Conservador":
            cfg["conservative_approach"] = True
            cfg["safety_margin"] = True
            print(f"[CONFIG] Conservador: enfoque conservador")
        
        print(f"[CONFIG] Configuración aplicada - Strategy: {cfg.get('strategy', 'default')}")
    else:
        print(f"[CONFIG] ADVERTENCIA: Perfil '{profile}' no encontrado, usando configuración por defecto")
    
    return cfg


def create_demand_signature(demand_matrix: np.ndarray) -> str:
    """Return a short hash representing the demand pattern."""
    normalized = demand_matrix / (demand_matrix.max() + 1e-8)
    return hashlib.md5(normalized.tobytes()).hexdigest()[:16]


def load_learning_history() -> dict:
    """Load adaptive learning history from disk if available."""
    try:
        if os.path.exists("learning_history.json"):
            with open("learning_history.json", "r") as fh:
                return json.load(fh)
    except Exception:
        pass
    return {}


def save_learning_history(history: dict) -> None:
    """Persist adaptive learning history to disk."""
    try:
        with open("learning_history.json", "w") as fh:
            json.dump(history, fh, indent=2)
    except Exception:
        pass


def get_adaptive_parameters(demand_signature: str, learning_history: dict) -> dict:
    """Return learned parameters for ``demand_signature`` or defaults."""
    if demand_signature in learning_history:
        learned = learning_history[demand_signature]
        best = min(learned.get("runs", []), key=lambda x: x.get("score", 0))
        return {
            "agent_limit_factor": best["params"]["agent_limit_factor"],
            "excess_penalty": best["params"]["excess_penalty"],
            "peak_bonus": best["params"]["peak_bonus"],
            "critical_bonus": best["params"]["critical_bonus"],
        }
    return {
        "agent_limit_factor": 22,
        "excess_penalty": 0.5,
        "peak_bonus": 1.5,
        "critical_bonus": 2.0,
    }


def update_learning_history(demand_signature: str, params: dict, results: dict, history: dict) -> dict:
    """Update ``history`` with new execution ``results`` for ``demand_signature``."""
    if demand_signature not in history:
        history[demand_signature] = {"runs": []}

    score = results["understaffing"] + results["overstaffing"] * 0.3
    history[demand_signature]["runs"].append(
        {
            "params": params,
            "score": score,
            "total_agents": results["total_agents"],
            "coverage": results["coverage_percentage"],
            "timestamp": time.time(),
        }
    )
    history[demand_signature]["runs"] = history[demand_signature]["runs"][-10:]
    return history


def get_optimal_break_time(start_hour: float, shift_duration: int, day: int, demand_day: np.ndarray,
                           *, cfg=None) -> float:
    """Return the best break time for the given day based on demand."""
    cfg = merge_config(cfg)
    break_earliest = start_hour + cfg["break_from_start"]
    break_latest = start_hour + shift_duration - cfg["break_from_end"]
    if break_latest <= break_earliest:
        return break_earliest
    options = []
    current = break_earliest
    while current <= break_latest:
        options.append(current)
        current += 0.5
    best = break_earliest
    min_impact = float("inf")
    for opt in options:
        hour = int(opt) % 24
        if hour < len(demand_day):
            impact = demand_day[hour]
            if impact < min_impact:
                min_impact = impact
                best = opt
    return best


def generate_shift_patterns(demand_matrix=None, *, top_k=100, cfg=None):
    """Generate patterns and keep only the ``top_k`` highest scoring ones.

    ``demand_matrix`` is the uncompressed 7x24 demand array.  It is packed
    internally once and the packed representation is reused for scoring.

    Patterns are evaluated against ``demand_matrix`` as they are produced and a
    min-heap is used to retain only the best ``top_k`` entries.  Each heap entry
    stores ``(score, name, pat_packed)`` where ``pat_packed`` is the flattened
    byte representation of the pattern.  The final result is returned as a list
    sorted by score in descending order.
    """
    cfg = merge_config(cfg)
    use_ft = cfg["use_ft"]
    use_pt = cfg["use_pt"]
    allow_8h = cfg["allow_8h"]
    allow_10h8 = cfg["allow_10h8"]
    allow_pt_4h = cfg["allow_pt_4h"]
    allow_pt_6h = cfg["allow_pt_6h"]
    allow_pt_5h = cfg["allow_pt_5h"]
    active_days = cfg["ACTIVE_DAYS"]

    first_hour = 6
    last_hour = 22
    demand_packed = None
    if demand_matrix is not None:
        analysis = analyze_demand_matrix(demand_matrix)
        first_hour = analysis["first_hour"]
        last_hour = analysis["last_hour"]
        demand_packed = np.packbits(demand_matrix > 0, axis=1).astype(np.uint8)

    start_hours = np.arange(max(6, first_hour), min(last_hour - 2, 20), 0.5)
    heap = []

    def push_pattern(name: str, pattern: np.ndarray) -> None:
        """Score and push ``pattern`` onto the heap, keeping only ``top_k``."""
        pat = pattern.astype(np.int8)
        score = score_pattern(pat, demand_packed) if demand_packed is not None else 0
        packed = pat.tobytes()
        entry = (score, name, packed)
        if len(heap) < top_k:
            heapq.heappush(heap, entry)
        else:
            if score > heap[0][0]:
                heapq.heapreplace(heap, entry)

    if use_ft and allow_8h:
        for start_hour in start_hours:
            for working_combo in combinations(active_days, min(6, len(active_days))):
                non_working = [d for d in active_days if d not in working_combo]
                for dso_day in non_working + [None]:
                    for brk in get_valid_break_times(start_hour, 8, cfg=cfg):
                        pattern = generate_weekly_pattern_with_break(
                            start_hour, 8, list(working_combo), dso_day, brk, cfg=cfg
                        )
                        suffix = f"_DSO{dso_day}" if dso_day is not None else ""
                        name = (
                            f"FT8_{start_hour:04.1f}_DAYS{''.join(map(str,working_combo))}_BRK{brk:04.1f}{suffix}"
                        )
                        push_pattern(name, pattern)

    if use_pt:
        if allow_pt_4h:
            for start_hour in start_hours:
                for num_days in [4, 5, 6]:
                    if num_days <= len(active_days):
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 4, list(combo))
                            name = f"PT4_{start_hour:04.1f}_DAYS{''.join(map(str,combo))}"
                            push_pattern(name, pattern)
        if allow_pt_6h:
            for start_hour in start_hours[::2]:
                for num_days in [4]:
                    if num_days <= len(active_days):
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 6, list(combo))
                            name = f"PT6_{start_hour:04.1f}_DAYS{''.join(map(str,combo))}"
                            push_pattern(name, pattern)
        if allow_pt_5h:
            for start_hour in start_hours[::2]:
                for num_days in [5]:
                    if num_days <= len(active_days):
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 5, list(combo))
                            name = f"PT5_{start_hour:04.1f}_DAYS{''.join(map(str,combo))}"
                            push_pattern(name, pattern)

    # Return heap contents sorted by score (highest first) and unpack patterns
    ordered = sorted(heap, reverse=True)
    return [
        (score, name, np.frombuffer(packed, dtype=np.int8))
        for score, name, packed in ordered
    ]


def get_valid_break_times(start_hour: float, duration: int, *, cfg=None) -> list:
    """Return a list of possible break start times for a shift."""
    cfg = merge_config(cfg)
    earliest = start_hour + cfg["break_from_start"]
    latest = start_hour + duration - cfg["break_from_end"] - 1
    valid = []
    current = earliest
    while current <= latest:
        if current % 0.5 == 0:
            valid.append(current)
        current += 0.5
    return valid[:7]


def generate_weekly_pattern_with_break(start_hour: float, duration: int, working_days: list, dso_day: int | None,
                                        break_start: float, break_len: int = 1, *, cfg=None) -> np.ndarray:
    """Generate a weekly pattern placing the break at ``break_start``."""
    cfg = merge_config(cfg)
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        if day == dso_day:
            continue
        for h in range(duration):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
        for b in range(int(break_len)):
            t = break_start + b
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 0
    return pattern.flatten()


def generate_weekly_pattern_advanced(start_hour: float, duration: int, working_days: list, break_position: float, *, cfg=None) -> np.ndarray:
    """Generate a weekly pattern with a break at ``break_position`` percentage of the shift."""
    cfg = merge_config(cfg)
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        for h in range(duration):
            hour_idx = int(start_hour + h) % 24
            if hour_idx < 24:
                pattern[day, hour_idx] = 1
        brk_offset = int(duration * break_position)
        brk_hour = int(start_hour + brk_offset) % 24
        if (
            brk_hour >= int(start_hour + cfg["break_from_start"])
            and brk_hour <= int(start_hour + duration - cfg["break_from_end"])
            and brk_hour < 24
        ):
            pattern[day, brk_hour] = 0
    return pattern.flatten()


def evaluate_solution_quality(coverage_matrix: np.ndarray, demand_matrix: np.ndarray) -> float:
    """Return a quality score for a coverage matrix (lower is better)."""
    total_demand = demand_matrix.sum()
    total_coverage = np.minimum(coverage_matrix, demand_matrix).sum()
    coverage_pct = (total_coverage / total_demand) * 100 if total_demand > 0 else 0
    excess = np.maximum(0, coverage_matrix - demand_matrix).sum()
    efficiency = total_coverage / (total_coverage + excess) if (total_coverage + excess) > 0 else 0
    return (100 - coverage_pct) + (excess * 0.1) + ((1 - efficiency) * 50)

# ---------------------------------------------------------------------------
# Pattern generation
# ---------------------------------------------------------------------------

def generate_weekly_pattern(start_hour, duration, working_days, dso_day=None, break_len=1, *, cfg=None):
    """Return a weekly pattern with a single break per day."""
    cfg = merge_config(cfg)
    break_from_start = cfg["break_from_start"]
    break_from_end = cfg["break_from_end"]
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        if day != dso_day:
            for h in range(duration):
                t = start_hour + h
                d_off, idx = divmod(int(t), 24)
                pattern[(day + d_off) % 7, idx] = 1
            break_start_idx = start_hour + break_from_start
            break_end_idx = start_hour + duration - break_from_end
            if int(break_start_idx) < int(break_end_idx):
                break_hour = int(break_start_idx) + (int(break_end_idx) - int(break_start_idx)) // 2
            else:
                break_hour = int(break_start_idx)
            for b in range(int(break_len)):
                t = break_hour + b
                d_off, idx = divmod(int(t), 24)
                pattern[(day + d_off) % 7, idx] = 0
    return pattern.flatten()


def generate_weekly_pattern_10h8(start_hour, working_days, eight_hour_day, break_len=1, *, cfg=None):
    """Generate a 10h pattern with one 8h day."""
    cfg = merge_config(cfg)
    break_from_start = cfg["break_from_start"]
    break_from_end = cfg["break_from_end"]
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        duration = 8 if day == eight_hour_day else 10
        for h in range(duration):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
        break_start_idx = start_hour + break_from_start
        break_end_idx = start_hour + duration - break_from_end
        if int(break_start_idx) < int(break_end_idx):
            break_hour = int(break_start_idx) + (int(break_end_idx) - int(break_start_idx)) // 2
        else:
            break_hour = int(break_start_idx)
        for b in range(int(break_len)):
            t = break_hour + b
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 0
    return pattern.flatten()


def generate_weekly_pattern_simple(start_hour, duration, working_days):
    """Simple pattern without breaks."""
    pattern = np.zeros((7, 24), dtype=np.int8)
    for day in working_days:
        for h in range(duration):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
    return pattern.flatten()


def generate_weekly_pattern_pt5(start_hour, working_days):
    """Five 5h days with the last reduced to 4h."""
    pattern = np.zeros((7, 24), dtype=np.int8)
    if not working_days:
        return pattern.flatten()
    four_hour_day = working_days[-1]
    for day in working_days:
        hours = 4 if day == four_hour_day else 5
        for h in range(hours):
            t = start_hour + h
            d_off, idx = divmod(int(t), 24)
            pattern[(day + d_off) % 7, idx] = 1
    return pattern.flatten()


def generate_shifts_coverage_corrected(*, max_patterns=None, batch_size=None, cfg=None):
    """Yield raw patterns respecting configuration flags."""
    cfg = merge_config(cfg)
    use_ft = cfg["use_ft"]
    use_pt = cfg["use_pt"]
    allow_8h = cfg["allow_8h"]
    allow_10h8 = cfg["allow_10h8"]
    allow_pt_4h = cfg["allow_pt_4h"]
    allow_pt_6h = cfg["allow_pt_6h"]
    allow_pt_5h = cfg["allow_pt_5h"]
    active_days = cfg["ACTIVE_DAYS"]

    shifts_coverage = {}
    seen_patterns = set()
    step = 0.5
    start_hours = [h for h in np.arange(0, 24, step) if h <= 23.5]
    
    for start_hour in start_hours:
        # Full Time 8 hours - 6 days of work
        if use_ft and allow_8h:
            for dso_day in active_days:
                working_days = [d for d in active_days if d != dso_day][:6]
                if len(working_days) >= 6 and 8 * len(working_days) <= 48:
                    pattern = generate_weekly_pattern(start_hour, 8, working_days, dso_day, cfg=cfg)
                    key = pattern.tobytes()
                    if key not in seen_patterns:
                        seen_patterns.add(key)
                        name = f"FT8_{start_hour:04.1f}_DSO{dso_day}"
                        shifts_coverage[name] = pattern
                        if batch_size and len(shifts_coverage) >= batch_size:
                            yield shifts_coverage
                            shifts_coverage = {}
        
        # Full Time 10h + 8h - 5 days of work
        if use_ft and allow_10h8:
            for dso_day in active_days:
                working_days = [d for d in active_days if d != dso_day][:5]
                if len(working_days) >= 5:
                    for eight_day in working_days:
                        pattern = generate_weekly_pattern_10h8(start_hour, working_days, eight_day, cfg=cfg)
                        key = pattern.tobytes()
                        if key not in seen_patterns:
                            seen_patterns.add(key)
                            name = f"FT10p8_{start_hour:04.1f}_DSO{dso_day}_8{eight_day}"
                            shifts_coverage[name] = pattern
                            if batch_size and len(shifts_coverage) >= batch_size:
                                yield shifts_coverage
                                shifts_coverage = {}
        
        # Part Time patterns
        if use_pt:
            # 4 hours - multiple day combinations
            if allow_pt_4h:
                for num_days in [4, 5, 6]:
                    if num_days <= len(active_days) and 4 * num_days <= 24:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 4, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT4_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
                                if batch_size and len(shifts_coverage) >= batch_size:
                                    yield shifts_coverage
                                    shifts_coverage = {}
            
            # 6 hours - 4 days (24h/week)
            if allow_pt_6h:
                for num_days in [4]:
                    if num_days <= len(active_days) and 6 * num_days <= 24:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_simple(start_hour, 6, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT6_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
                                if batch_size and len(shifts_coverage) >= batch_size:
                                    yield shifts_coverage
                                    shifts_coverage = {}
            
            # 5 hours - 5 days (~25h/week)
            if allow_pt_5h:
                for num_days in [5]:
                    if num_days <= len(active_days) and 5 * num_days <= 25:
                        for combo in combinations(active_days, num_days):
                            pattern = generate_weekly_pattern_pt5(start_hour, list(combo))
                            key = pattern.tobytes()
                            if key not in seen_patterns:
                                seen_patterns.add(key)
                                name = f"PT5_{start_hour:04.1f}_DAYS{''.join(map(str, combo))}"
                                shifts_coverage[name] = pattern
                                if batch_size and len(shifts_coverage) >= batch_size:
                                    yield shifts_coverage
                                    shifts_coverage = {}
    
    if shifts_coverage:
        yield shifts_coverage


def generate_shifts_coverage_optimized(
    demand_matrix,
    *,
    max_patterns=None,
    batch_size=2000,
    quality_threshold=0,
    cfg=None,
    demand_packed=None,
):
    """Yield scored pattern batches.

    Logic mirrors ``generate_shifts_coverage_optimized`` in ``legacy/app1.py``
    but omits any Streamlit UI elements.

    Configuration keys:
        K (int): maximum number of high-scoring patterns kept when the
            ``JEAN Personalizado`` profile is active.
    """
    cfg = merge_config(cfg)
    profile = cfg.get('optimization_profile', 'Equilibrado (Recomendado)')
    
    # JEAN Personalizado: usar load_shift_patterns si hay JSON personalizado
    if profile == 'JEAN Personalizado' and cfg.get('custom_shifts_json'):
        slot_minutes = int(cfg.get('slot_duration_minutes', 30))
        start_hours = [h for h in np.arange(0, 24, slot_minutes / 60) if h <= 23.5]
        patterns = load_shift_patterns(
            cfg['custom_shifts_json'],
            start_hours=start_hours,
            break_from_start=cfg.get('break_from_start', DEFAULT_CONFIG['break_from_start']),
            break_from_end=cfg.get('break_from_end', DEFAULT_CONFIG['break_from_end']),
            slot_duration_minutes=slot_minutes,
            demand_matrix=demand_matrix,
            keep_percentage=cfg.get('keep_percentage', 0.3),
            peak_bonus=cfg.get('peak_bonus', 1.5),
            critical_bonus=cfg.get('critical_bonus', 2.0),
            efficiency_bonus=cfg.get('efficiency_bonus', 1.0),
            max_patterns=cfg.get('max_patterns')
        )
        
        # Filtrar por tipo de contrato
        if not cfg.get('use_ft', True):
            patterns = {k: v for k, v in patterns.items() if not k.startswith('FT')}
        if not cfg.get('use_pt', True):
            patterns = {k: v for k, v in patterns.items() if not k.startswith('PT')}
        
        k = cfg.get('K', 1000)
        heap = []
        for name, pat in patterns.items():
            score = score_pattern(pat, demand_matrix)
            heapq.heappush(heap, (score, name, pat))
            if len(heap) > k:
                heapq.heappop(heap)
        yield {name: pat for _, name, pat in heap}
        return
    
    # Generación estándar para otros perfiles
    selected = 0
    seen = set()
    inner = generate_shifts_coverage_corrected(batch_size=batch_size, cfg=cfg)
    for raw_batch in inner:
        batch = {}
        for name, pat in raw_batch.items():
            key = hashlib.md5(pat).digest()
            if key in seen:
                continue
            seen.add(key)
            if score_pattern(pat, demand_matrix) >= quality_threshold:
                batch[name] = pat
                selected += 1
                if max_patterns is not None and selected >= max_patterns:
                    break
        if batch:
            yield batch
            gc.collect()
        if max_patterns is not None and selected >= max_patterns:
            break


@single_model
def optimize_jean_search(shifts_coverage, demand_matrix, *, cfg=None, target_coverage=98.0, max_iterations=5, job_id=None):
    """Búsqueda iterativa JEAN EXACTA del legacy original."""
    import time
    start_time = time.time()
    max_time = 120  # Máximo 2 minutos para JEAN
    
    cfg = merge_config(cfg)
    original_factor = cfg["agent_limit_factor"]
    
    best_assignments = {}
    best_method = ""
    best_score = float("inf")
    best_coverage = 0
    
    print(f"[JEAN] Iniciando búsqueda iterativa JEAN")
    
    # Función para actualizar progreso
    def update_progress(info):
        if job_id:
            try:
                from .extensions import scheduler as store
                store.update_progress(job_id, info)
            except Exception:
                pass
    
    # Secuencia de factores EXACTA del legacy Streamlit - REDUCIDA para evitar timeout
    factor_sequence = [30, 20, 15, 10]  # Secuencia más corta
    factor_sequence = [f for f in factor_sequence if f <= original_factor]
    
    if not factor_sequence:
        factor_sequence = [original_factor]
    
    print(f"[JEAN] Secuencia de factores: {factor_sequence}")
    
    # Implementar EXACTAMENTE la lógica del legacy Streamlit
    for iteration, factor in enumerate(factor_sequence[:max_iterations]):
        # Verificar timeout
        if time.time() - start_time > max_time:
            print(f"[JEAN] Timeout alcanzado ({max_time}s), terminando")
            break
            
        print(f"[JEAN] Iteración {iteration + 1}: factor {factor}")
        update_progress({"jean_iteration": f"Factor {factor} ({iteration + 1}/{len(factor_sequence)})"})
        
        # Actualizar configuración temporal EXACTA del legacy
        temp_cfg = cfg.copy()
        temp_cfg["agent_limit_factor"] = factor
        temp_cfg["solver_time"] = 30  # Timeout agresivo por iteración
        
        try:
            assignments, method = optimize_with_precision_targeting(shifts_coverage, demand_matrix, cfg=temp_cfg, job_id=job_id)
            results = analyze_results(assignments, shifts_coverage, demand_matrix)
            
            if results:
                cov = results["coverage_percentage"]
                score = results["overstaffing"] + results["understaffing"]
                print(f"[JEAN] Factor {factor}: cobertura {cov:.1f}%, score {score:.1f}")
                
                if cov >= target_coverage:
                    if score < best_score or not best_assignments:
                        best_assignments, best_method = assignments, f"JEAN_SEARCH_F{factor}"
                        best_score = score
                        best_coverage = cov
                        print(f"[JEAN] Nueva mejor solución: score {score:.1f}")
                    # En JEAN, continuar buscando mejores scores incluso si se alcanza cobertura
                elif cov > best_coverage:
                    best_assignments, best_method, best_coverage = assignments, f"JEAN_SEARCH_F{factor}", cov
                    best_score = score
                    print(f"[JEAN] Mejor cobertura parcial: {cov:.1f}%")
        except Exception as e:
            print(f"[JEAN] Error en iteración {iteration + 1}: {e}")
            # En caso de error, continuar con el siguiente factor
            pass
    
    # Si no hay resultados, usar greedy como fallback
    if not best_assignments:
        print(f"[JEAN] Sin resultados, usando greedy como fallback")
        best_assignments, best_method = optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
    
    elapsed = time.time() - start_time
    print(f"[JEAN] Búsqueda completada en {elapsed:.1f}s: mejor score {best_score:.1f}, cobertura {best_coverage:.1f}%")
    return best_assignments, best_method


@single_model
def optimize_perfect_coverage(shifts_coverage, demand_matrix, *, cfg=None):
    """Optimización para cobertura perfecta (100%)."""
    cfg = merge_config(cfg)
    profile = cfg.get("optimization_profile", "")
    
    print(f"[PERFECT] Iniciando optimización de cobertura perfecta: {profile}")
    
    if not PULP_AVAILABLE:
        return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg)
    
    try:
        shifts_list = list(shifts_coverage.keys())
        if not shifts_list:
            return {}, "NO_SHIFTS"
        
        prob = pl.LpProblem("Perfect_Coverage", pl.LpMinimize)
        
        # Variables con límites según el perfil
        total_demand = demand_matrix.sum()
        if profile == "100% Cobertura Total":
            max_per_shift = max(20, int(total_demand / 3))  # Más generoso
        else:
            max_per_shift = max(15, int(total_demand / cfg["agent_limit_factor"]))
        
        shift_vars = {}
        for shift in shifts_list:
            shift_vars[shift] = pl.LpVariable(f"shift_{shift}", 0, max_per_shift, pl.LpInteger)
        
        # Variables de déficit y exceso
        deficit_vars = {}
        excess_vars = {}
        hours = demand_matrix.shape[1]
        patterns_unpacked = {
            s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
            for s, p in shifts_coverage.items()
        }
        
        for day in range(7):
            for hour in range(hours):
                deficit_vars[(day, hour)] = pl.LpVariable(f"deficit_{day}_{hour}", 0, None)
                excess_vars[(day, hour)] = pl.LpVariable(f"excess_{day}_{hour}", 0, None)
        
        # Función objetivo según el perfil
        total_deficit = pl.lpSum([deficit_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_excess = pl.lpSum([excess_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_agents = pl.lpSum([shift_vars[shift] for shift in shifts_list])
        
        if profile == "100% Exacto":
            # Prohibir cualquier déficit o exceso
            prob += total_deficit * 1000000 + total_excess * 1000000 + total_agents * 1
            # Restricciones estrictas
            prob += total_deficit == 0
            prob += total_excess == 0
        elif profile == "100% Cobertura Eficiente":
            # Minimizar exceso pero permitir cobertura completa
            prob += total_deficit * 100000 + total_excess * cfg["excess_penalty"] * 1000 + total_agents * 1
            prob += total_deficit == 0  # Sin déficit
            prob += total_excess <= total_demand * cfg.get("max_excess_ratio", 0.02)
        elif profile == "100% Cobertura Total":
            # Cobertura completa sin restricciones de exceso
            prob += total_deficit * 100000 + total_excess * cfg["excess_penalty"] + total_agents * 0.1
            prob += total_deficit == 0  # Sin déficit
        else:
            # Cobertura Perfecta - balance
            prob += total_deficit * 50000 + total_excess * cfg["excess_penalty"] * 100 + total_agents * 1
        
        # Restricciones de cobertura
        for day in range(7):
            for hour in range(hours):
                coverage = pl.lpSum([
                    shift_vars[shift] * patterns_unpacked[shift][day, hour]
                    for shift in shifts_list
                ])
                demand = demand_matrix[day, hour]
                
                prob += coverage + deficit_vars[(day, hour)] >= demand
                prob += coverage - excess_vars[(day, hour)] <= demand
        
        # Resolver con tiempo extendido
        solver_time = cfg.get("solver_time", 600)
        solver = pl.PULP_CBC_CMD(
            msg=cfg.get("solver_msg", 0),
            timeLimit=solver_time,
            threads=cfg["solver_threads"],
            gapRel=0.001 if profile == "100% Exacto" else 0.01
        )
        
        prob.solve(solver)
        
        assignments = {}
        if prob.status == pl.LpStatusOptimal:
            for shift in shifts_list:
                val = int(shift_vars[shift].varValue or 0)
                if val > 0:
                    assignments[shift] = val
            return assignments, f"PERFECT_{profile.upper().replace(' ', '_')}"
        else:
            print(f"[PERFECT] Status no óptimo: {prob.status}, usando fallback")
            return optimize_with_precision_targeting(shifts_coverage, demand_matrix, cfg=cfg)
    
    except Exception as e:
        print(f"[PERFECT] Error: {e}")
        return optimize_with_precision_targeting(shifts_coverage, demand_matrix, cfg=cfg)


@single_model
def optimize_with_precision_targeting(shifts_coverage, demand_matrix, *, cfg=None, job_id=None):
    """Optimización ultra-precisa EXACTA del legacy original."""
    cfg = merge_config(cfg)
    
    print(f"[PRECISION] Iniciando optimización de precisión")
    
    if not PULP_AVAILABLE:
        return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
    
    try:
        shifts_list = list(shifts_coverage.keys())
        if not shifts_list:
            return {}, "NO_SHIFTS"
        
        prob = pl.LpProblem("Precision_Scheduling", pl.LpMinimize)
        
        # Variables con límites dinámicos EXACTOS del legacy
        total_demand = demand_matrix.sum()
        peak_demand = demand_matrix.max()
        max_per_shift = max(20, int(total_demand / cfg["agent_limit_factor"]))
        
        shift_vars = {}
        for shift in shifts_list:
            shift_vars[shift] = pl.LpVariable(f"shift_{shift}", 0, max_per_shift, pl.LpInteger)
        
        # Variables de déficit y exceso
        deficit_vars = {}
        excess_vars = {}
        hours = demand_matrix.shape[1]
        patterns_unpacked = {
            s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
            for s, p in shifts_coverage.items()
        }
        
        for day in range(7):
            for hour in range(hours):
                deficit_vars[(day, hour)] = pl.LpVariable(f"deficit_{day}_{hour}", 0, None)
                excess_vars[(day, hour)] = pl.LpVariable(f"excess_{day}_{hour}", 0, None)
        
        # Análisis de patrones críticos EXACTO del legacy
        daily_totals = demand_matrix.sum(axis=1)
        hourly_totals = demand_matrix.sum(axis=0)
        critical_days = np.argsort(daily_totals)[-2:] if len(daily_totals) > 1 else [np.argmax(daily_totals)]
        peak_threshold = np.percentile(hourly_totals[hourly_totals > 0], 75) if np.any(hourly_totals > 0) else 0
        peak_hours = np.where(hourly_totals >= peak_threshold)[0]
        
        # Función objetivo EXACTA del legacy
        total_deficit = pl.lpSum([deficit_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_excess = pl.lpSum([excess_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_agents = pl.lpSum([shift_vars[shift] for shift in shifts_list])
        
        # Bonificaciones por días críticos y horas pico EXACTAS del legacy
        critical_bonus_value = 0
        peak_bonus_value = 0
        
        # Días críticos
        for critical_day in critical_days:
            if critical_day < 7:
                for hour in range(hours):
                    if demand_matrix[critical_day, hour] > 0:
                        critical_bonus_value -= deficit_vars[(critical_day, hour)] * cfg["critical_bonus"]
        
        # Horas pico
        for hour in peak_hours:
            if hour < hours:
                for day in range(7):
                    if demand_matrix[day, hour] > 0:
                        peak_bonus_value -= deficit_vars[(day, hour)] * cfg["peak_bonus"]
        
        # Función objetivo EXACTA del legacy
        prob += (total_deficit * 1000 + 
                 total_excess * cfg["excess_penalty"] + 
                 total_agents * 0.1 + 
                 critical_bonus_value + 
                 peak_bonus_value)
        
        # Restricciones de cobertura EXACTAS del legacy
        for day in range(7):
            for hour in range(hours):
                coverage = pl.lpSum([
                    shift_vars[shift] * patterns_unpacked[shift][day, hour]
                    for shift in shifts_list
                ])
                demand = demand_matrix[day, hour]
                
                prob += coverage + deficit_vars[(day, hour)] >= demand
                prob += coverage - excess_vars[(day, hour)] <= demand
        
        # Restricciones adicionales según perfil EXACTAS del legacy
        profile = cfg.get("optimization_profile", "")
        if cfg["excess_penalty"] > 5:  # Perfiles estrictos como "100% Exacto"
            prob += total_excess <= demand_matrix.sum() * 0.02
        elif cfg["excess_penalty"] > 2:
            prob += total_excess <= demand_matrix.sum() * 0.05
        
        # Límite dinámico de agentes EXACTO del legacy
        dynamic_agent_limit = max(
            int(total_demand / max(1, cfg["agent_limit_factor"])),
            int(peak_demand * 1.1),
        )
        prob += total_agents <= dynamic_agent_limit
        
        # Resolver EXACTO del legacy con timeout más agresivo
        solver_time = min(cfg.get("solver_time", 240), 60)  # Máximo 60 segundos
        solver = pl.PULP_CBC_CMD(
            msg=cfg.get("solver_msg", 0), 
            timeLimit=solver_time,
            threads=min(cfg["solver_threads"], 4),  # Limitar threads
            gapRel=0.05  # Permitir 5% de gap para terminar más rápido
        )
        prob.solve(solver)
        
        # Extraer solución EXACTA del legacy
        assignments = {}
        if prob.status == pl.LpStatusOptimal:
            for shift in shifts_list:
                value = int(shift_vars[shift].varValue or 0)
                if value > 0:
                    assignments[shift] = value
            method = "PRECISION_TARGETING"
        elif prob.status == pl.LpStatusInfeasible:
            print(f"[PRECISION] Problema infactible, relajando restricciones")
            return optimize_with_relaxed_constraints(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
        else:
            print(f"[PRECISION] Solver status: {prob.status}, usando fallback")
            return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
        
        return assignments, method
        
    except Exception as e:
        print(f"[PRECISION] Error: {e}")
        return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)


@single_model
def optimize_with_relaxed_constraints(shifts_coverage, demand_matrix, *, cfg=None, job_id=None):
    """Optimización con restricciones relajadas EXACTA del legacy"""
    cfg = merge_config(cfg)
    
    if not PULP_AVAILABLE:
        return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
    
    try:
        shifts_list = list(shifts_coverage.keys())
        if not shifts_list:
            return {}, "NO_SHIFTS"
        
        prob = pl.LpProblem("Relaxed_Scheduling", pl.LpMinimize)
        
        # Variables con límites muy generosos
        total_demand = demand_matrix.sum()
        max_per_shift = max(20, int(total_demand / 5))
        
        shift_vars = {}
        for shift in shifts_list:
            shift_vars[shift] = pl.LpVariable(f"shift_{shift}", 0, max_per_shift, pl.LpInteger)
        
        # Solo variables de déficit (sin restricciones de exceso)
        deficit_vars = {}
        hours = demand_matrix.shape[1]
        patterns_unpacked = {
            s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
            for s, p in shifts_coverage.items()
        }
        
        for day in range(7):
            for hour in range(hours):
                deficit_vars[(day, hour)] = pl.LpVariable(f"deficit_{day}_{hour}", 0, None)
        
        # Objetivo simple: minimizar déficit
        total_deficit = pl.lpSum([deficit_vars[(day, hour)] for day in range(7) for hour in range(hours)])
        total_agents = pl.lpSum([shift_vars[shift] for shift in shifts_list])
        
        prob += total_deficit * 1000 + total_agents * 0.1
        
        # Solo restricciones básicas de cobertura
        for day in range(7):
            for hour in range(hours):
                coverage = pl.lpSum([
                    shift_vars[shift] * patterns_unpacked[shift][day, hour]
                    for shift in shifts_list
                ])
                demand = demand_matrix[day, hour]
                prob += coverage + deficit_vars[(day, hour)] >= demand
        
        # Límite muy generoso de agentes
        prob += total_agents <= int(total_demand / 3)
        
        # Resolver con configuración básica y timeout agresivo
        prob.solve(pl.PULP_CBC_CMD(
            msg=0, 
            timeLimit=min(cfg.get("solver_time", 120), 30),
            threads=2,
            gapRel=0.2
        ))
        
        assignments = {}
        if prob.status == pl.LpStatusOptimal:
            for shift in shifts_list:
                value = int(shift_vars[shift].varValue or 0)
                if value > 0:
                    assignments[shift] = value
            return assignments, "RELAXED_CONSTRAINTS"
        else:
            return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
            
    except Exception as e:
        print(f"[RELAXED] Error: {e}")
        return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)


def optimize_ft_then_pt_strategy(shifts_coverage, demand_matrix, *, cfg=None, job_id=None):
    """Estrategia 2 fases: FT sin exceso, luego PT para completar."""
    print("[FT_PT] Iniciando estrategia 2 fases")
    cfg = merge_config(cfg)
    
    ft_shifts = {k: v for k, v in shifts_coverage.items() if k.startswith('FT')}
    pt_shifts = {k: v for k, v in shifts_coverage.items() if k.startswith('PT')}

    # Reducir el problema: quedarnos con los mejores patrones según demanda
    def _top_by_score(shifts_dict, max_n):
        if max_n is None or len(shifts_dict) <= max_n:
            return shifts_dict
        scored = [(name, pat, score_pattern(pat, demand_matrix)) for name, pat in shifts_dict.items()]
        scored.sort(key=lambda x: x[2], reverse=True)
        keep = dict((name, pat) for name, pat, _ in scored[:max_n])
        return keep

    ft_limit = cfg.get("ft_max_patterns")
    pt_limit = cfg.get("pt_max_patterns")
    if ft_limit:
        update_job_progress(job_id, {"stage": f"Filtrando FT (top {ft_limit})"})
        ft_shifts = _top_by_score(ft_shifts, int(ft_limit))
    if pt_limit:
        update_job_progress(job_id, {"stage": f"Filtrando PT (top {pt_limit})"})
        pt_shifts = _top_by_score(pt_shifts, int(pt_limit))
    
    print(f"[FT_PT] Fase 1: {len(ft_shifts)} turnos FT")
    update_job_progress(job_id, {"stage": "Resolviendo FT"})
    ft_assignments = optimize_ft_no_excess(ft_shifts, demand_matrix, cfg=cfg)
    update_job_progress(job_id, {"stage": "FT resuelto", "ft_turnos": len(ft_assignments)})
    
    # Calcular cobertura FT
    ft_coverage = np.zeros_like(demand_matrix)
    for name, count in ft_assignments.items():
        pattern = ft_shifts[name].reshape(7, demand_matrix.shape[1]) if len(ft_shifts[name]) == 7 * demand_matrix.shape[1] else ft_shifts[name].reshape(7, -1)[:, :demand_matrix.shape[1]]
        ft_coverage += pattern * count
    
    print(f"[FT_PT] Fase 2: {len(pt_shifts)} turnos PT")
    remaining_demand = np.maximum(0, demand_matrix - ft_coverage)
    update_job_progress(job_id, {"stage": "Resolviendo PT"})

    # Publicar resultado parcial (solo FT) para que la UI tenga algo visible
    try:
        if job_id is not None and ft_assignments:
            from .extensions import scheduler as store
            partial_metrics = analyze_results(ft_assignments, ft_shifts, demand_matrix, ft_coverage)
            partial_result = {
                "assignments": ft_assignments,
                "metrics": partial_metrics,
                "status": "PARTIAL_FT",
                "pulp_results": {
                    "assignments": ft_assignments,
                    "metrics": partial_metrics,
                    "status": "FT_ONLY",
                    "heatmaps": {},
                },
                "greedy_results": {"assignments": {}, "metrics": None, "status": "NOT_EXECUTED", "heatmaps": {}},
                "effective_config": merge_config(cfg),
            }
            store.mark_finished(job_id, partial_result, None, None)
            # Guardar snapshot parcial en disco
            try:
                path = os.path.join(tempfile.gettempdir(), f"scheduler_result_{job_id}.json")
                with open(path, "w", encoding="utf-8") as fh:
                    json.dump(partial_result, fh, ensure_ascii=False)
            except Exception:
                pass
            update_job_progress(job_id, {"stage": "Publicado resultado parcial (FT)"})
    except Exception:
        pass
    pt_assignments = optimize_pt_complete(pt_shifts, remaining_demand, cfg=cfg)
    update_job_progress(job_id, {"stage": "PT resuelto", "pt_turnos": len(pt_assignments)})
    
    final_assignments = {**ft_assignments, **pt_assignments}
    print(f"[FT_PT] Completado: {len(final_assignments)} turnos asignados")
    update_job_progress(job_id, {"stage": "FT+PT completado", "turnos": len(final_assignments)})
    return final_assignments, "FT_NO_EXCESS_THEN_PT"


@single_model
def optimize_ft_no_excess(ft_shifts, demand_matrix, *, cfg=None):
    """Linear program focusing on full-time coverage only."""
    cfg = merge_config(cfg)
    agent_limit_factor = cfg["agent_limit_factor"]
    solver_time = cfg.get("solver_time")
    if not ft_shifts:
        return {}
    prob = pl.LpProblem("FT_No_Excess", pl.LpMinimize)
    max_ft_per_shift = max(10, int(demand_matrix.sum() / agent_limit_factor))
    ft_vars = {s: pl.LpVariable(f"ft_{s}", 0, max_ft_per_shift, pl.LpInteger) for s in ft_shifts}
    deficit_vars = {}
    hours = demand_matrix.shape[1]
    patterns_unpacked = {
        s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
        for s, p in ft_shifts.items()
    }
    for d in range(7):
        for h in range(hours):
            deficit_vars[(d, h)] = pl.LpVariable(f"ft_deficit_{d}_{h}", 0, None)
    total_deficit = pl.lpSum(deficit_vars.values())
    total_ft_agents = pl.lpSum(ft_vars.values())
    prob += total_deficit * 1000 + total_ft_agents * 1
    for d in range(7):
        for h in range(hours):
            coverage = pl.lpSum(ft_vars[s] * patterns_unpacked[s][d, h] for s in ft_shifts)
            demand = demand_matrix[d, h]
            prob += coverage + deficit_vars[(d, h)] >= demand
            prob += coverage <= demand
    solver_kwargs = {
        "msg": cfg.get("solver_msg", 1), 
        "threads": min(cfg["solver_threads"], 2),
        "gapRel": 0.1
    }
    if solver_time is not None:
        solver_kwargs["timeLimit"] = min(solver_time // 2, 30)
    else:
        solver_kwargs["timeLimit"] = 30
    prob.solve(pl.PULP_CBC_CMD(**solver_kwargs))
    assignments = {}
    if prob.status == pl.LpStatusOptimal:
        for s in ft_shifts:
            val = int(ft_vars[s].varValue or 0)
            if val > 0:
                assignments[s] = val
    return assignments


@single_model
def optimize_pt_complete(pt_shifts, remaining_demand, *, cfg=None):
    """Solve for part-time assignments covering ``remaining_demand``."""
    cfg = merge_config(cfg)
    agent_limit_factor = cfg["agent_limit_factor"]
    excess_penalty = cfg["excess_penalty"]
    solver_time = cfg.get("solver_time")
    optimization_profile = cfg["optimization_profile"]
    if not pt_shifts or remaining_demand.sum() == 0:
        return {}
    prob = pl.LpProblem("PT_Complete", pl.LpMinimize)
    max_pt_per_shift = max(10, int(remaining_demand.sum() / max(1, agent_limit_factor)))
    pt_vars = {s: pl.LpVariable(f"pt_{s}", 0, max_pt_per_shift, pl.LpInteger) for s in pt_shifts}
    deficit_vars = {}
    excess_vars = {}
    hours = remaining_demand.shape[1]
    patterns_unpacked = {
        s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
        for s, p in pt_shifts.items()
    }
    for d in range(7):
        for h in range(hours):
            deficit_vars[(d, h)] = pl.LpVariable(f"pt_deficit_{d}_{h}", 0, None)
            excess_vars[(d, h)] = pl.LpVariable(f"pt_excess_{d}_{h}", 0, None)
    total_deficit = pl.lpSum(deficit_vars.values())
    total_excess = pl.lpSum(excess_vars.values())
    total_pt_agents = pl.lpSum(pt_vars.values())
    prob += total_deficit * 1000 + total_excess * (excess_penalty * 20) + total_pt_agents * 1
    if optimization_profile in ("JEAN", "JEAN Personalizado"):
        prob += total_excess == 0
    for d in range(7):
        for h in range(hours):
            coverage = pl.lpSum(pt_vars[s] * patterns_unpacked[s][d, h] for s in pt_shifts)
            demand = remaining_demand[d, h]
            prob += coverage + deficit_vars[(d, h)] >= demand
            prob += coverage - excess_vars[(d, h)] <= demand
    solver_kwargs = {
        "msg": cfg.get("solver_msg", 1), 
        "threads": min(cfg["solver_threads"], 2),
        "gapRel": 0.1
    }
    if solver_time is not None:
        solver_kwargs["timeLimit"] = min(solver_time // 2, 30)
    else:
        solver_kwargs["timeLimit"] = 30
    prob.solve(pl.PULP_CBC_CMD(**solver_kwargs))
    assignments = {}
    if prob.status == pl.LpStatusOptimal:
        for s in pt_shifts:
            val = int(pt_vars[s].varValue or 0)
            if val > 0:
                assignments[s] = val
    return assignments


def optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, *, cfg=None, job_id=None):
    """Algoritmo greedy mejorado EXACTO del legacy Streamlit."""
    print("[GREEDY] Iniciando algoritmo greedy mejorado")
    cfg = merge_config(cfg)
    agent_limit_factor = cfg["agent_limit_factor"]
    excess_penalty = cfg["excess_penalty"]
    peak_bonus = cfg["peak_bonus"]
    critical_bonus = cfg["critical_bonus"]

    shifts_list = list(shifts_coverage.keys())
    assignments = {}
    current_coverage = np.zeros_like(demand_matrix, dtype=float)
    
    # Límite de agentes EXACTO del legacy
    max_agents = max(100, int(demand_matrix.sum() / max(1, agent_limit_factor - 5)))
    
    print(f"[GREEDY] Procesando {len(shifts_list)} turnos, max {max_agents} agentes")
    update_job_progress(job_id, {"greedy_iteration": f"0/{max_agents}"})

    # Análisis de patrones críticos EXACTO del legacy
    daily_totals = demand_matrix.sum(axis=1)
    hourly_totals = demand_matrix.sum(axis=0)
    critical_days = (
        np.argsort(daily_totals)[-2:]
        if daily_totals.size > 1
        else [int(np.argmax(daily_totals))]
    )
    peak_threshold = (
        np.percentile(hourly_totals[hourly_totals > 0], 75)
        if np.any(hourly_totals > 0)
        else 0
    )
    peak_hours = np.where(hourly_totals >= peak_threshold)[0]
    
    print(f"[GREEDY] Días críticos: {critical_days}, Horas pico: {peak_hours}")

    for iteration in range(max_agents):
        if iteration % 20 == 0:
            print(f"[GREEDY] Iteración {iteration}/{max_agents}")
            update_job_progress(job_id, {"greedy_iteration": f"{iteration}/{max_agents}"})
            
        best_shift = None
        best_score = -float("inf")
        best_pattern = None

        for shift_name in shifts_list:
            try:
                # Obtener patrón del turno EXACTO del legacy
                slots_per_day = len(shifts_coverage[shift_name]) // 7
                base_pattern = np.array(shifts_coverage[shift_name]).reshape(7, slots_per_day)
                
                # Ajustar dimensiones si es necesario
                if slots_per_day != demand_matrix.shape[1]:
                    pattern = np.zeros((7, demand_matrix.shape[1]))
                    cols_to_copy = min(slots_per_day, demand_matrix.shape[1])
                    pattern[:, :cols_to_copy] = base_pattern[:, :cols_to_copy]
                else:
                    pattern = base_pattern
                
                new_coverage = current_coverage + pattern
                
                # Cálculo de score EXACTO del legacy
                current_deficit = np.maximum(0, demand_matrix - current_coverage)
                new_deficit = np.maximum(0, demand_matrix - new_coverage)
                deficit_reduction = np.sum(current_deficit - new_deficit)
                
                # Penalización inteligente de exceso EXACTA del legacy
                current_excess = np.maximum(0, current_coverage - demand_matrix)
                new_excess = np.maximum(0, new_coverage - demand_matrix)
                
                smart_excess_penalty = 0
                for day in range(7):
                    for hour in range(demand_matrix.shape[1]):
                        if demand_matrix[day, hour] == 0 and new_excess[day, hour] > current_excess[day, hour]:
                            smart_excess_penalty += 1000  # Penalización extrema
                        elif demand_matrix[day, hour] <= 2:
                            smart_excess_penalty += (new_excess[day, hour] - current_excess[day, hour]) * excess_penalty * 10
                        else:
                            smart_excess_penalty += (new_excess[day, hour] - current_excess[day, hour]) * excess_penalty
                
                # Bonificaciones por patrones críticos EXACTAS del legacy
                critical_bonus_score = 0
                for critical_day in critical_days:
                    if critical_day < 7:
                        day_improvement = np.sum(current_deficit[critical_day] - new_deficit[critical_day])
                        critical_bonus_score += day_improvement * critical_bonus * 2
                
                peak_bonus_score = 0
                for hour in peak_hours:
                    if hour < demand_matrix.shape[1]:
                        hour_improvement = np.sum(current_deficit[:, hour] - new_deficit[:, hour])
                        peak_bonus_score += hour_improvement * peak_bonus * 2
                
                # Score final EXACTO del legacy
                score = (deficit_reduction * 100 + 
                        critical_bonus_score + 
                        peak_bonus_score - 
                        smart_excess_penalty)
                
                if score > best_score:
                    best_score = score
                    best_shift = shift_name
                    best_pattern = pattern
                    
            except Exception as e:
                print(f"[GREEDY] Error procesando {shift_name}: {e}")
                continue
        
        # Criterio de parada EXACTO del legacy
        if best_shift is None or best_score <= 0.5:
            print(f"[GREEDY] Parada en iteración {iteration}, mejor score: {best_score}")
            break
        
        # Aplicar mejor turno
        if best_shift not in assignments:
            assignments[best_shift] = 0
        assignments[best_shift] += 1
        current_coverage += best_pattern
        
        # Verificar si se completó la cobertura
        remaining_deficit = np.sum(np.maximum(0, demand_matrix - current_coverage))
        if remaining_deficit == 0:
            print(f"[GREEDY] Cobertura completa en iteración {iteration}")
            break
    
    total_agents = sum(assignments.values())
    print(f"[GREEDY] Completado: {total_agents} agentes asignados en {len(assignments)} turnos")
    return assignments, "GREEDY_ENHANCED"


def _generate_partial_result(assignments, patterns, demand_matrix, cfg):
    """Genera resultado parcial para mostrar mientras otros algoritmos siguen ejecutándose."""
    try:
        metrics = analyze_results(assignments, patterns, demand_matrix)
        return {
            "assignments": assignments,
            "metrics": metrics,
            "status": "PARTIAL_GREEDY",
            "partial": True
        }
    except Exception as e:
        print(f"[PARTIAL] Error generando resultado parcial: {e}")
        return None


@single_model
def solve_with_pulp(demand_matrix, patterns, config):
    """Solve assignment problem using PuLP.

    Parameters
    ----------
    demand_matrix : numpy.ndarray
        Required coverage for each day/hour.
    patterns : dict[str, numpy.ndarray]
        Mapping of shift name to a flattened weekly pattern.
    config : dict
        Configuration overrides.

    Returns
    -------
    dict
        Dictionary with assignments, coverage matrix, total agents and status.
    """
    if not PULP_AVAILABLE:
        raise RuntimeError("PuLP is not available")

    cfg = merge_config(config)
    hours = demand_matrix.shape[1]
    shifts = list(patterns.keys())
    patterns_unpacked = {
        s: p.reshape(7, hours) if len(p) == 7 * hours else p.reshape(7, -1)[:, :hours]
        for s, p in patterns.items()
    }

    prob = pl.LpProblem("schedule_with_pulp", pl.LpMinimize)
    max_per_shift = int(demand_matrix.max() * cfg["agent_limit_factor"])

    shift_vars = {
        s: pl.LpVariable(f"shift_{i}", lowBound=0, upBound=max_per_shift, cat=pl.LpInteger)
        for i, s in enumerate(shifts)
    }
    deficit = {
        (d, h): pl.LpVariable(f"def_{d}_{h}", lowBound=0, cat=pl.LpContinuous)
        for d in range(7)
        for h in range(hours)
    }
    excess = {
        (d, h): pl.LpVariable(f"exc_{d}_{h}", lowBound=0, cat=pl.LpContinuous)
        for d in range(7)
        for h in range(hours)
    }

    total_agents = pl.lpSum(shift_vars.values())
    total_deficit = pl.lpSum(deficit.values())
    total_excess = pl.lpSum(excess.values())
    prob += total_agents + 1000 * total_deficit + cfg["excess_penalty"] * total_excess

    for d in range(7):
        for h in range(hours):
            coverage_expr = pl.lpSum(
                shift_vars[s] * patterns_unpacked[s][d, h] for s in shifts
            )
            demand = demand_matrix[d, h]
            prob += coverage_expr + deficit[(d, h)] >= demand
            prob += coverage_expr - excess[(d, h)] <= demand

    solver_kwargs = {"msg": cfg.get("solver_msg", 1), "threads": 1, "timeLimit": cfg.get("solver_time", 300)}
    solver = pl.PULP_CBC_CMD(**solver_kwargs)
    status = prob.solve(solver)

    assignments = {
        s: int(pl.value(v))
        for s, v in shift_vars.items()
        if pl.value(v) and pl.value(v) > 0.5
    }

    coverage = np.zeros_like(demand_matrix, dtype=float)
    for s, count in assignments.items():
        coverage += patterns_unpacked[s] * count

    total_agents_val = int(sum(assignments.values()))
    solver_status = pl.LpStatus.get(prob.status, str(prob.status))
    return {
        "assignments": assignments,
        "coverage_matrix": coverage,
        "total_agents": total_agents_val,
        "status": solver_status,
    }


def optimize_schedule_iterative_legacy(shifts_coverage, demand_matrix, *, cfg=None, job_id=None):
    """Función principal EXACTA del legacy Streamlit con estrategia FT primero + PT después."""
    cfg = merge_config(cfg)
    
    if PULP_AVAILABLE:
        profile = cfg.get("optimization_profile", "Equilibrado (Recomendado)")
        
        if profile == "JEAN":
            print("[LEGACY] Búsqueda JEAN: cobertura sin exceso")
            return optimize_jean_search(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=cfg.get("TARGET_COVERAGE", 98.0))

        if profile == "JEAN Personalizado":
            if cfg.get("use_ft") and cfg.get("use_pt"):
                print("[LEGACY] Estrategia 2 Fases: FT sin exceso -> PT para completar")
                assignments, method = optimize_ft_then_pt_strategy(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)

                results = analyze_results(assignments, shifts_coverage, demand_matrix)
                if results:
                    cov = results["coverage_percentage"]
                    score = results["overstaffing"] + results["understaffing"]
                    target_coverage = cfg.get("TARGET_COVERAGE", 98.0)
                    if cov < target_coverage or score > 0:
                        print("[LEGACY] Refinando con búsqueda JEAN")
                        assignments, method = optimize_jean_search(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=target_coverage)

                return assignments, method
            else:
                print("[LEGACY] Búsqueda JEAN: cobertura sin exceso")
                return optimize_jean_search(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=cfg.get("TARGET_COVERAGE", 98.0))

        if cfg.get("use_ft") and cfg.get("use_pt"):
            print("[LEGACY] Estrategia 2 Fases: FT sin exceso -> PT para completar")
            return optimize_ft_then_pt_strategy(shifts_coverage, demand_matrix, cfg=cfg, job_id=job_id)
        else:
            print("[LEGACY] Modo Precisión: Optimización directa")
            return optimize_with_precision_targeting(shifts_coverage, demand_matrix, cfg=cfg)
    else:
        print("[LEGACY] Solver Básico: Greedy mejorado")
        return optimize_schedule_greedy_enhanced(shifts_coverage, demand_matrix, cfg=cfg)


def optimize_jean_personalizado_legacy(shifts_coverage, demand_matrix, *, cfg=None, job_id=None):
    """JEAN Personalizado EXACTO del legacy con carga de patrones JSON."""
    cfg = merge_config(cfg)
    
    print(f"[JEAN_CUSTOM] Iniciando JEAN Personalizado legacy")
    
    # Si hay configuración JSON personalizada, cargar patrones
    if cfg.get("custom_shifts_json"):
        try:
            slot_minutes = int(cfg.get("slot_duration_minutes", 30))
            start_hours = [h for h in np.arange(0, 24, slot_minutes / 60) if h <= 23.5]
            
            custom_patterns = load_shift_patterns(
                cfg["custom_shifts_json"],
                start_hours=start_hours,
                break_from_start=cfg.get("break_from_start", DEFAULT_CONFIG["break_from_start"]),
                break_from_end=cfg.get("break_from_end", DEFAULT_CONFIG["break_from_end"]),
                slot_duration_minutes=slot_minutes,
                demand_matrix=demand_matrix,
                keep_percentage=cfg.get("keep_percentage", 0.3),
                peak_bonus=cfg.get("peak_bonus", 1.5),
                critical_bonus=cfg.get("critical_bonus", 2.0),
                efficiency_bonus=cfg.get("efficiency_bonus", 1.0),
                max_patterns=cfg.get("max_patterns")
            )
            
            if custom_patterns:
                # Filtrar por tipo de contrato
                if not cfg.get("use_ft", True):
                    custom_patterns = {k: v for k, v in custom_patterns.items() if not k.startswith("FT")}
                if not cfg.get("use_pt", True):
                    custom_patterns = {k: v for k, v in custom_patterns.items() if not k.startswith("PT")}
                
                shifts_coverage = custom_patterns
                print(f"[JEAN_CUSTOM] Cargados {len(custom_patterns)} patrones personalizados")
        except Exception as e:
            print(f"[JEAN_CUSTOM] Error cargando patrones personalizados: {e}")
    
    # Estrategia 2 fases si se usan FT y PT
    if cfg.get("use_ft") and cfg.get("use_pt"):
        print(f"[JEAN_CUSTOM] Usando estrategia 2 fases FT->PT")
        assignments, status = optimize_ft_then_pt_strategy(shifts_coverage, demand_matrix, cfg=cfg)
        
        # Verificar si necesita refinamiento con búsqueda JEAN
        results = analyze_results(assignments, shifts_coverage, demand_matrix)
        if results:
            coverage = results["coverage_percentage"]
            score = results["overstaffing"] + results["understaffing"]
            target = cfg.get("TARGET_COVERAGE", 98.0)
            
            if coverage < target or score > 0:
                print(f"[JEAN_CUSTOM] Refinando con búsqueda JEAN (cov: {coverage:.1f}%, score: {score:.1f})")
                refined_assignments, refined_status = optimize_jean_search(
                    shifts_coverage, demand_matrix, cfg=cfg, target_coverage=target, job_id=job_id
                )
                if refined_assignments:
                    return refined_assignments, f"JEAN_CUSTOM_REFINED_{refined_status}"
        
        return assignments, f"JEAN_CUSTOM_{status}"
    else:
        # Usar búsqueda JEAN estándar
        return optimize_jean_search(shifts_coverage, demand_matrix, cfg=cfg, target_coverage=cfg.get("TARGET_COVERAGE", 98.0), job_id=job_id)


def solve_in_chunks_optimized_legacy(shifts_coverage, demand_matrix, *, cfg=None):
    """Solve EXACTO del legacy Streamlit usando chunks ordenados por score."""
    scored = []
    seen = set()
    for name, pat in shifts_coverage.items():
        key = hashlib.md5(pat).digest()
        if key in seen:
            continue
        seen.add(key)
        scored.append((name, pat, score_pattern(pat, demand_matrix)))

    scored.sort(key=lambda x: x[2], reverse=True)

    assignments_total = {}
    coverage = np.zeros_like(demand_matrix)
    idx = 0
    base_chunk_size = 10000
    
    while idx < len(scored):
        chunk_size = adaptive_chunk_size(base_chunk_size)
        chunk_dict = {name: pat for name, pat, _ in scored[idx: idx + chunk_size]}
        remaining = np.maximum(0, demand_matrix - coverage)
        if not np.any(remaining):
            break
        
        # Usar la función de optimización según el perfil
        assigns, _ = optimize_schedule_iterative_legacy(chunk_dict, remaining, cfg=cfg)
        
        for name, val in assigns.items():
            assignments_total[name] = assignments_total.get(name, 0) + val
            slots = len(chunk_dict[name]) // 7
            pat_matrix = chunk_dict[name].reshape(7, slots)
            if slots != demand_matrix.shape[1]:
                # Ajustar dimensiones si es necesario
                adjusted_pattern = np.zeros((7, demand_matrix.shape[1]))
                cols_to_copy = min(slots, demand_matrix.shape[1])
                adjusted_pattern[:, :cols_to_copy] = pat_matrix[:, :cols_to_copy]
                pat_matrix = adjusted_pattern
            coverage += pat_matrix * val
        
        idx += chunk_size
        gc.collect()
        emergency_cleanup()
        
        if not np.any(np.maximum(0, demand_matrix - coverage)):
            break
    
    return assignments_total


def analyze_results(assignments, shifts_coverage, demand_matrix, coverage_matrix=None):
    """Compute coverage metrics from solved assignments - EXACTO del legacy."""
    if not assignments:
        return None

    compute_coverage = coverage_matrix is None
    if compute_coverage:
        # Usar las mismas dimensiones que demand_matrix
        coverage_matrix = np.zeros_like(demand_matrix, dtype=np.int16)
    
    total_agents = 0
    ft_agents = 0
    pt_agents = 0
    
    for shift_name, count in assignments.items():
        total_agents += count
        if shift_name.startswith('FT'):
            ft_agents += count
        else:
            pt_agents += count
            
        if compute_coverage and shift_name in shifts_coverage:
            weekly_pattern = shifts_coverage[shift_name]
            target_shape = demand_matrix.shape
            
            # Conversión de dimensiones EXACTA del legacy
            if len(weekly_pattern) == target_shape[0] * target_shape[1]:
                pattern_matrix = weekly_pattern.reshape(target_shape)
            else:
                # Reshape temporal y ajuste
                slots_per_day = len(weekly_pattern) // 7
                pattern_temp = weekly_pattern.reshape(7, slots_per_day)
                pattern_matrix = np.zeros(target_shape)
                
                if slots_per_day == target_shape[1]:
                    # Dimensiones coinciden
                    pattern_matrix = pattern_temp
                else:
                    # Ajustar dimensiones copiando lo que se pueda
                    cols_to_copy = min(slots_per_day, target_shape[1])
                    pattern_matrix[:, :cols_to_copy] = pattern_temp[:, :cols_to_copy]
            
            coverage_matrix += pattern_matrix * count

    # Cálculo de métricas EXACTO del legacy
    total_demand = demand_matrix.sum()
    total_covered = np.minimum(coverage_matrix, demand_matrix).sum()
    
    # Coverage percentage: demanda cubierta / demanda total (ponderado por horas)
    coverage_percentage = (
        (total_covered / total_demand) * 100 if total_demand > 0 else 0
    )
    
    diff_matrix = coverage_matrix - demand_matrix
    overstaffing = np.sum(diff_matrix[diff_matrix > 0])
    understaffing = np.sum(np.abs(diff_matrix[diff_matrix < 0]))
    
    return {
        'total_coverage': coverage_matrix,
        'total_agents': total_agents,
        'ft_agents': ft_agents,
        'pt_agents': pt_agents,
        'coverage_percentage': coverage_percentage,
        'overstaffing': overstaffing,
        'understaffing': understaffing,
        'diff_matrix': diff_matrix,
    }


def _extract_start_hour(name: str) -> float:
    """Best-effort extraction of the start hour from a shift name."""
    for part in name.split('_'):
        if '.' in part and part.replace('.', '').isdigit():
            try:
                return float(part)
            except ValueError:
                continue
    return 0.0


def export_detailed_schedule(assignments, shifts_coverage):
    """Return schedule data in Excel and CSV formats.

    The Excel workbook is created using ``openpyxl`` in write-only mode and
    populated row by row.  A CSV version of the detailed schedule is generated
    simultaneously using the :mod:`csv` module.  The function returns a tuple
    ``(excel_bytes, csv_bytes)``.  If ``assignments`` is empty, ``(None, None)``
    is returned.
    """

    if not assignments:
        return None, None

    from openpyxl import Workbook
    from collections import defaultdict

    DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

    wb = Workbook(write_only=True)
    ws_details = wb.create_sheet('Horarios_Semanales')
    ws_summary = wb.create_sheet('Resumen_Agentes')
    ws_shifts = wb.create_sheet('Turnos_Asignados')

    ws_details.append(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])
    ws_summary.append(['Agente', 'Turno', 'Dias_Trabajo'])
    ws_shifts.append(['Turno', 'Agentes'])

    csv_buffer = StringIO()
    csv_writer = csv.writer(csv_buffer)
    csv_writer.writerow(['Agente', 'Dia', 'Horario', 'Break', 'Turno', 'Tipo'])

    summary_counts = defaultdict(int)

    agent_id = 1
    for shift_name, count in assignments.items():
        weekly_pattern = shifts_coverage[shift_name]
        slots_per_day = 24  # Always use 24 hours per day
        pattern_matrix = weekly_pattern.reshape(7, slots_per_day) if len(weekly_pattern) == 7 * slots_per_day else weekly_pattern.reshape(7, -1)[:, :slots_per_day]
        parts = shift_name.split('_')
        start_hour = _extract_start_hour(shift_name)
        if shift_name.startswith('FT10p8'):
            shift_type = 'FT'
            shift_duration = 10
            total_hours = shift_duration + 1
        elif shift_name.startswith('FT'):
            shift_type = 'FT'
            if '_Variado_' in shift_name or len(parts) > 4:
                shift_duration = int(pattern_matrix.sum(axis=1).max())
            else:
                try:
                    if len(parts[0]) > 2 and parts[0][2:].isdigit():
                        shift_duration = int(parts[0][2:])
                    else:
                        shift_duration = int(pattern_matrix.sum(axis=1).max())
                except (ValueError, IndexError):
                    shift_duration = int(pattern_matrix.sum(axis=1).max())
            total_hours = shift_duration + 1
        elif shift_name.startswith('PT'):
            shift_type = 'PT'
            if '_Variado_' in shift_name or len(parts) > 4:
                shift_duration = int(pattern_matrix.sum(axis=1).max())
            else:
                try:
                    if len(parts[0]) > 2 and parts[0][2:].isdigit():
                        shift_duration = int(parts[0][2:])
                    else:
                        shift_duration = int(pattern_matrix.sum(axis=1).max())
                except (ValueError, IndexError):
                    shift_duration = int(pattern_matrix.sum(axis=1).max())
            total_hours = shift_duration
        else:
            shift_type = 'FT'
            shift_duration = 8
            total_hours = 9

        for _ in range(count):
            for day in range(7):
                day_pattern = pattern_matrix[day]
                work_hours = np.where(day_pattern == 1)[0]
                if len(work_hours) > 0:
                    start_idx = int(start_hour)
                    if shift_name.startswith('PT') or shift_name.startswith('FT10p8'):
                        end_idx = (int(work_hours[-1]) + 1) % 24
                        next_day = end_idx <= start_idx
                        horario = f"{start_idx:02d}:00-{end_idx:02d}:00" + ("+1" if next_day else "")
                    else:
                        end_idx = int(work_hours[-1]) + 1
                        next_day = end_idx <= start_idx
                        horario = f"{start_idx:02d}:00-{end_idx % 24:02d}:00" + ("+1" if next_day else "")
                    if shift_name.startswith('PT'):
                        break_time = ""
                    elif shift_name.startswith('FT10p8'):
                        all_expected = set(range(int(start_hour), int(start_hour + total_hours)))
                        actual_hours = set(work_hours)
                        break_hours = all_expected - actual_hours
                        if break_hours:
                            break_hour = min(break_hours) % 24
                            break_end = (break_hour + 1) % 24 or 24
                            break_time = f"{break_hour:02d}:00-{break_end:02d}:00"
                        else:
                            break_time = ""
                    else:
                        expected = list(range(start_idx, end_idx))
                        if next_day:
                            expected = list(range(start_idx, 24)) + list(range(0, end_idx % 24))
                        break_hours = set(expected) - set(work_hours)
                        if break_hours:
                            break_hour = min(break_hours)
                            break_time = f"{break_hour % 24:02d}:00-{((break_hour + 1) % 24) or 24:02d}:00"
                        else:
                            break_time = ""
                    row = [f"AGT_{agent_id:03d}", DAYS[day], horario, break_time, shift_name, shift_type]
                else:
                    row = [f"AGT_{agent_id:03d}", DAYS[day], "DSO", "", shift_name, "DSO"]
                ws_details.append(row)
                csv_writer.writerow(row)
                summary_counts[(agent_id, shift_name)] += 1
            agent_id += 1

    for (agent_idx, shift_name), cnt in summary_counts.items():
        ws_summary.append([f"AGT_{agent_idx:03d}", shift_name, cnt])

    for shift, count in assignments.items():
        ws_shifts.append([shift, count])

    excel_io = BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()

    csv_bytes = csv_buffer.getvalue().encode("utf-8")
    return excel_bytes, csv_bytes


def run_optimization(file_stream, config=None):
    """Main optimization function matching legacy interface."""
    return run_complete_optimization(file_stream, config, generate_charts=False)


def update_job_progress(job_id, progress_info):
    """Actualizar progreso de un job específico."""
    if job_id:
        try:
            from .extensions import scheduler as store
            store.update_progress(job_id, progress_info)
        except Exception as e:
            print(f"[PROGRESS] Error actualizando progreso: {e}")


def run_complete_optimization(file_stream, config=None, generate_charts=False, job_id=None):
    """Run the full optimization pipeline matching legacy behavior."""
    print("[SCHEDULER] Iniciando run_complete_optimization")
    print(f"[SCHEDULER] Config recibido: {config}")

    if job_id:
        active_jobs[job_id] = current_thread()
        update_job_progress(job_id, {"stage": "Iniciando optimización"})

    try:
        cfg = apply_configuration(config)

        print(f"[MEM] Antes de carga de demanda: {monitor_memory_usage():.1f}%")
        print("[SCHEDULER] Leyendo archivo Excel...")
        import pandas as pd
        df = pd.read_excel(file_stream)
        print("[SCHEDULER] Archivo Excel leído correctamente")

        print("[SCHEDULER] Procesando matriz de demanda...")
        demand_matrix = load_demand_matrix_from_df(df)
        analysis = analyze_demand_matrix(demand_matrix)
        print("[SCHEDULER] Matriz de demanda procesada")
        print(f"[MEM] Después de procesar demanda: {monitor_memory_usage():.1f}%")

        print(f"[MEM] Antes de generación de patrones: {monitor_memory_usage():.1f}%")
        print("[SCHEDULER] Generando patrones de turnos...")
        patterns = {}
        if cfg.get("optimization_profile") == "JEAN Personalizado" and cfg.get("custom_shifts_json"):
            # Solo usar load_shift_patterns si hay un JSON personalizado
            slot_minutes = int(cfg.get("slot_duration_minutes", 30))
            start_hours = [h for h in np.arange(0, 24, slot_minutes / 60) if h <= 23.5]
            patterns = load_shift_patterns(
                cfg["custom_shifts_json"],
                start_hours=start_hours,
                break_from_start=cfg.get(
                    "break_from_start", DEFAULT_CONFIG["break_from_start"]
                ),
                break_from_end=cfg.get(
                    "break_from_end", DEFAULT_CONFIG["break_from_end"]
                ),
                slot_duration_minutes=slot_minutes,
                demand_matrix=demand_matrix,
                keep_percentage=cfg.get("keep_percentage", 0.3),
                peak_bonus=cfg.get("peak_bonus", 1.5),
                critical_bonus=cfg.get("critical_bonus", 2.0),
                efficiency_bonus=cfg.get("efficiency_bonus", 1.0),
                max_patterns=cfg.get("max_patterns"),
            )
            if not cfg.get("use_ft", True):
                patterns = {k: v for k, v in patterns.items() if not k.startswith("FT")}
            if not cfg.get("use_pt", True):
                patterns = {k: v for k, v in patterns.items() if not k.startswith("PT")}
        else:
            for batch in generate_shifts_coverage_optimized(
                demand_matrix,
                max_patterns=cfg.get("max_patterns"),
                batch_size=cfg.get("batch_size", 2000),
                quality_threshold=cfg.get("quality_threshold", 0),
                cfg=cfg,
            ):
                patterns.update(batch)
                if cfg.get("max_patterns") and len(patterns) >= cfg["max_patterns"]:
                    break
        print("[SCHEDULER] Patrones generados")
        print(f"[MEM] Después de generación de patrones: {monitor_memory_usage():.1f}%")

        print("[SCHEDULER] Iniciando optimizacion...")
        print(f"[MEM] Antes de resolver: {monitor_memory_usage():.1f}%")
        
        # Usar lógica EXACTA del legacy Streamlit
        print("[SCHEDULER] Ejecutando optimización con lógica legacy...")
        
        profile = cfg.get("optimization_profile", "Equilibrado (Recomendado)")
        
        # LÓGICA EXACTA DEL LEGACY STREAMLIT
        if profile == "JEAN":
            print("[SCHEDULER] Ejecutando búsqueda JEAN")
            assignments, status = optimize_jean_search(
                patterns, demand_matrix, cfg=cfg, 
                target_coverage=cfg.get("TARGET_COVERAGE", 98.0),
                max_iterations=cfg.get("search_iterations", 5),
                job_id=job_id
            )
        elif profile == "JEAN Personalizado":
            print("[SCHEDULER] Ejecutando JEAN Personalizado")
            # Si hay FT y PT, usar estrategia 2 fases + refinamiento JEAN
            if cfg.get("use_ft") and cfg.get("use_pt"):
                print("[SCHEDULER] Estrategia 2 fases FT->PT + refinamiento JEAN")
                assignments, status = optimize_ft_then_pt_strategy(patterns, demand_matrix, cfg=cfg, job_id=job_id)
                
                # Verificar si necesita refinamiento
                results = analyze_results(assignments, patterns, demand_matrix)
                if results:
                    coverage = results["coverage_percentage"]
                    score = results["overstaffing"] + results["understaffing"]
                    target = cfg.get("TARGET_COVERAGE", 98.0)
                    
                    if coverage < target or score > 0:
                        print(f"[SCHEDULER] Refinando con JEAN (cov: {coverage:.1f}%, score: {score:.1f})")
                        refined_assignments, refined_status = optimize_jean_search(
                            patterns, demand_matrix, cfg=cfg, 
                            target_coverage=target,
                            job_id=job_id
                        )
                        if refined_assignments:
                            assignments, status = refined_assignments, f"JEAN_CUSTOM_REFINED_{refined_status}"
            else:
                # Solo un tipo de contrato, usar JEAN directo
                assignments, status = optimize_jean_search(
                    patterns, demand_matrix, cfg=cfg, 
                    target_coverage=cfg.get("TARGET_COVERAGE", 98.0),
                    job_id=job_id
                )
        elif profile == "Aprendizaje Adaptativo":
            print("[SCHEDULER] Ejecutando Aprendizaje Adaptativo")
            # Aplicar parámetros adaptativos
            adaptive_params = get_adaptive_params(demand_matrix, cfg.get("TARGET_COVERAGE", 98.0))
            temp_cfg = cfg.copy()
            temp_cfg.update(adaptive_params)
            assignments, status = optimize_schedule_iterative_legacy(patterns, demand_matrix, cfg=temp_cfg)
        else:
            print(f"[SCHEDULER] Ejecutando perfil estándar: {profile}")
            assignments, status = optimize_schedule_iterative_legacy(patterns, demand_matrix, cfg=cfg, job_id=job_id)
        
        total_agents = sum(assignments.values()) if assignments else 0
        
        # Simular resultados de PuLP y Greedy para compatibilidad
        pulp_assignments, pulp_status = assignments, status
        greedy_assignments, greedy_status = {}, "NOT_EXECUTED"
        
        print(f"[SCHEDULER] Optimización legacy completada: {len(assignments)} turnos, {total_agents} agentes")
        update_job_progress(job_id, {"stage": "Cálculo de métricas"})
        
        warning_msg = None
        
        # Calcular coverage matrix - usar dimensiones de la demanda
        coverage_matrix = np.zeros_like(demand_matrix)
        if assignments:
            for name, count in assignments.items():
                if name in patterns:
                    try:
                        pattern_flat = patterns[name]
                        # Convertir a las dimensiones de demand_matrix
                        target_shape = demand_matrix.shape
                        
                        if len(pattern_flat) == target_shape[0] * target_shape[1]:
                            # Dimensiones coinciden exactamente
                            pattern = pattern_flat.reshape(target_shape)
                        else:
                            # Convertir a matriz temporal y luego ajustar
                            pattern_temp = pattern_flat.reshape(7, -1)
                            pattern = np.zeros(target_shape)
                            
                            if pattern_temp.shape[1] == target_shape[1]:
                                # Mismo número de columnas
                                pattern = pattern_temp
                            elif pattern_temp.shape[1] == 192 and target_shape[1] == 24:
                                # Convertir de 192 slots a 24 horas
                                for day in range(7):
                                    for hour in range(24):
                                        start_slot = hour * 8
                                        end_slot = start_slot + 8
                                        pattern[day, hour] = np.mean(pattern_temp[day, start_slot:end_slot])
                            elif pattern_temp.shape[1] == 24 and target_shape[1] == 192:
                                # Convertir de 24 horas a 192 slots
                                for day in range(7):
                                    for hour in range(24):
                                        start_slot = hour * 8
                                        end_slot = start_slot + 8
                                        pattern[day, start_slot:end_slot] = pattern_temp[day, hour]
                            else:
                                # Fallback: copiar lo que se pueda
                                cols_to_copy = min(pattern_temp.shape[1], target_shape[1])
                                pattern[:, :cols_to_copy] = pattern_temp[:, :cols_to_copy]
                        
                        coverage_matrix += pattern * count
                    except Exception as e:
                        print(f"[SCHEDULER] Error procesando patrón {name}: {e}")
                        continue
        
        print(f"[MEM] Después de resolver: {monitor_memory_usage():.1f}%")
        print(f"[OPTIMIZER] Status: {status}")
        print(f"[OPTIMIZER] Total agents: {total_agents}")
        print("\u2705 [SCHEDULER] Optimización completada")
        
        # Agregar resultados de ambos algoritmos
        pulp_metrics = analyze_results(pulp_assignments, patterns, demand_matrix) if pulp_assignments else None
        greedy_metrics = analyze_results(greedy_assignments, patterns, demand_matrix) if greedy_assignments else None

        print(f"[MEM] Antes de exportar resultados: {monitor_memory_usage():.1f}%")
        update_job_progress(job_id, {"stage": "Exportación (opcional)"})
        metrics = analyze_results(assignments, patterns, demand_matrix, coverage_matrix)
        # Exportacin de archivos opcional para evitar bloquear la entrega de resultados
        export_files = bool(cfg.get("export_files", False))
        if export_files:
            excel_bytes, csv_bytes = export_detailed_schedule(assignments, patterns)
        else:
            excel_bytes, csv_bytes = None, None

        # Save learning data if adaptive mode
        if cfg.get("optimization_profile") == "Aprendizaje Adaptativo" and metrics:
            save_execution_result(
                demand_matrix, cfg, metrics["coverage_percentage"], 
                total_agents, 0  # execution_time placeholder
            )

        heatmaps = {}
        pulp_heatmaps = {}
        greedy_heatmaps = {}
        
        if generate_charts:
            # Gráficas generales (demanda)
            demand_maps = generate_all_heatmaps(demand_matrix)
            for key, fig in demand_maps.items():
                if key == "demand":  # Solo la gráfica de demanda es común
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    fig.savefig(tmp.name, format="png", bbox_inches="tight")
                    tmp.flush()
                    tmp.close()
                    filename = os.path.basename(tmp.name)
                    heatmaps[key] = f"/heatmap/{filename}"
                plt.close(fig)
            
            # Gráficas específicas de PuLP
            if pulp_metrics and pulp_assignments:
                pulp_maps = generate_all_heatmaps(
                    demand_matrix,
                    pulp_metrics.get("total_coverage"),
                    pulp_metrics.get("diff_matrix"),
                )
                for key, fig in pulp_maps.items():
                    if key != "demand":  # Excluir demanda ya que es común
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        fig.savefig(tmp.name, format="png", bbox_inches="tight")
                        tmp.flush()
                        tmp.close()
                        filename = os.path.basename(tmp.name)
                        pulp_heatmaps[key] = f"/heatmap/{filename}"
                    plt.close(fig)
            
            # Gráficas específicas de Greedy
            if greedy_metrics and greedy_assignments:
                greedy_maps = generate_all_heatmaps(
                    demand_matrix,
                    greedy_metrics.get("total_coverage"),
                    greedy_metrics.get("diff_matrix"),
                )
                for key, fig in greedy_maps.items():
                    if key != "demand":  # Excluir demanda ya que es común
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        fig.savefig(tmp.name, format="png", bbox_inches="tight")
                        tmp.flush()
                        tmp.close()
                        filename = os.path.basename(tmp.name)
                        greedy_heatmaps[key] = f"/heatmap/{filename}"
                    plt.close(fig)
            
            plt.close("all")
        print(f"[MEM] Después de exportar resultados: {monitor_memory_usage():.1f}%")

        print("[SCHEDULER] Preparando resultados...")

        def _convert(obj):
            """Recursively convert numpy arrays within ``obj`` to Python lists."""
            if isinstance(obj, np.ndarray):
                return obj.tolist()
            if isinstance(obj, dict):
                return {k: _convert(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [_convert(v) for v in obj]
            return obj

        result = {
            "analysis": _convert(analysis),
            "assignments": assignments,
            "metrics": _convert(metrics),
            "heatmaps": heatmaps,
            "status": status,
            "pulp_results": {
                "assignments": pulp_assignments,
                "metrics": _convert(pulp_metrics),
                "status": pulp_status,
                "heatmaps": pulp_heatmaps
            },
            "greedy_results": {
                "assignments": greedy_assignments,
                "metrics": _convert(greedy_metrics),
                "status": greedy_status,
                "heatmaps": greedy_heatmaps
            }
        }
        if warning_msg:
            result["message"] = warning_msg
        result["effective_config"] = _convert(cfg)
        print("[SCHEDULER] Resultados preparados - RETORNANDO")
        update_job_progress(job_id, {"stage": "Resultados listos"})

        # Publicar resultados en el store inmediatamente para que la UI los vea
        # incluso si hay pasos posteriores opcionales. El worker lo volverá a
        # marcar como finished al retornar, lo cual es idempotente.
        try:
            if job_id is not None:
                from .extensions import scheduler as store
                store.mark_finished(job_id, result, excel_bytes, csv_bytes)
                # Persistir respaldo en disco para que la vista pueda leerlo
                # incluso si hubiera algún problema con el store en memoria.
                try:
                    path = os.path.join(tempfile.gettempdir(), f"scheduler_result_{job_id}.json")
                    with open(path, "w", encoding="utf-8") as fh:
                        json.dump(result, fh, ensure_ascii=False)
                except Exception:
                    pass
        except Exception:
            pass

        # Callback removido - no hay variables use_greedy ni results_queue definidas

        return result, excel_bytes, csv_bytes

    except Exception as e:
        print(f"[SCHEDULER] ERROR CRÍTICO: {str(e)}")
        return {"error": str(e)}, None, None
    finally:
        if job_id:
            active_jobs.pop(job_id, None)
