import csv
import io
import os
import uuid
import tempfile
from openpyxl import Workbook


def process_file(file_storage):
    """Process uploaded CSV/XLSX file and return basic summary and download bytes."""
    filename = file_storage.filename or ""
    ext = os.path.splitext(filename.lower())[1]
    file_storage.seek(0)
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(file_storage)
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    else:
        text = file_storage.read().decode("utf-8")
        rows = list(csv.reader(io.StringIO(text)))
    file_storage.seek(0)

    header = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    summary_html = (
        "<table class='table table-striped'><tr><th>Filas</th><td>%d</td></tr></table>"
        % len(data_rows)
    )

    # create csv bytes
    csv_buffer = io.StringIO()
    csv.writer(csv_buffer).writerows(rows)
    csv_bytes = csv_buffer.getvalue().encode("utf-8")

    # create xlsx bytes
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_bytes = xlsx_buffer.getvalue()

    # placeholder heatmap: generate empty temp file
    heatmap_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4().hex}.png")
    with open(heatmap_path, "wb") as f:
        f.write(b"")

    result = {"tables": {"summary": summary_html}, "heatmap": heatmap_path}
    return result, csv_bytes, xlsx_bytes
